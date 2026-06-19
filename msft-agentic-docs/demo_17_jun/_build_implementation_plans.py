"""
Build two VendorPulse implementation-plan workbooks from the CR-9 EGB scope.

  1. VendorPulse_Implementation_Plan_Client.xlsx   - manager-friendly, low-detail,
     every dependency and date still shown.
  2. VendorPulse_Implementation_Plan_Detailed.xlsx - full task breakdown plus
     supporting sheets (pre-mob dependency checklist, quality gates, risk register,
     acceptance criteria).

Format follows the existing "Implementation Plan (1).xlsx":
    header fill  = FF00B0F0   phase-banner fill = FFFFC000
    dates        = d-mmm-yy   font = Aptos Narrow

Architecture baseline (demo_17_jun, latest):
    - Self-implemented agent layer (BaseAgent tool-calling).  MAF SDK is NOT used.
    - Provider-abstracted LLM: Anthropic Claude recommended (enterprise zero-retention /
      no-training / EU-residency DPA); Azure OpenAI (gpt-4o-class) config-only alternative.
    - Azure App Service (Linux containers) behind Front Door + WAF, Postgres Flexible Server,
      Key Vault + Managed Identity, App Insights + Log Analytics, all in Shell's tenant.

T-0 (Day 1) = Mon 22-Jun-2026.  Four Mon-Fri development/deployment weeks.
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------- styling
HEADER_FILL = PatternFill("solid", fgColor="FF00B0F0")
PHASE_FILL  = PatternFill("solid", fgColor="FFFFC000")
GATE_FILL   = PatternFill("solid", fgColor="FFFFF2CC")   # pale gold - quality gates
BLOCK_FILL  = PatternFill("solid", fgColor="FFFCE4D6")   # pale orange - hard Day-1 blocker
LATER_FILL  = PatternFill("solid", fgColor="FFE2EFDA")   # pale green  - needed later
TITLE_FILL  = PatternFill("solid", fgColor="FF1F4E78")   # deep blue   - title band
SUB_FILL    = PatternFill("solid", fgColor="FFDDEBF7")   # pale blue   - sub heading

FONT        = "Aptos Narrow"
HEAD_FONT   = Font(name=FONT, bold=True, size=11)
PHASE_FONT  = Font(name=FONT, bold=True, size=12)
TITLE_FONT  = Font(name=FONT, bold=True, size=16, color="FFFFFFFF")
SUBT_FONT   = Font(name=FONT, bold=True, size=10, color="FF1F4E78")
CELL_FONT   = Font(name=FONT, size=10)
BOLD_FONT   = Font(name=FONT, size=10, bold=True)

thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR_TOP  = Alignment(horizontal="center", vertical="top", wrap_text=True)
CTR_MID  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_MID = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATEFMT  = "d-mmm-yy"

T0 = dt.date(2026, 6, 22)          # Day 1
def d(y, m, day): return dt.date(y, m, day)

# Key calendar anchors (Mon-Fri)
W1 = (d(2026,6,22), d(2026,6,26))   # Days 1-5
W2 = (d(2026,6,29), d(2026,7,3))    # Days 6-10
W3 = (d(2026,7,6),  d(2026,7,10))   # Days 11-15
W4 = (d(2026,7,13), d(2026,7,17))   # Days 16-20

# --------------------------------------------------------------------------- helpers
def style_row(ws, r, c0, ncol, fill=None, font=CELL_FONT, align=WRAP_TOP):
    for i in range(ncol):
        cell = ws.cell(row=r, column=c0 + i)
        cell.border = BORDER
        cell.font = font
        cell.alignment = align
        if fill:
            cell.fill = fill

def banner(ws, r, c0, ncol, text, fill=PHASE_FILL, font=PHASE_FONT, height=20):
    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + ncol - 1)
    cell = ws.cell(row=r, column=c0, value=text)
    cell.fill = fill; cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i in range(ncol):
        ws.cell(row=r, column=c0 + i).border = BORDER
    ws.row_dimensions[r].height = height

def header(ws, r, c0, cols):
    for i, h in enumerate(cols):
        cell = ws.cell(row=r, column=c0 + i, value=h)
        cell.fill = HEADER_FILL; cell.font = HEAD_FONT
        cell.alignment = CTR_MID; cell.border = BORDER
    ws.row_dimensions[r].height = 30

def write_data_row(ws, r, c0, values, fill=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=r, column=c0 + i, value=v)
        cell.border = BORDER
        cell.font = CELL_FONT
        if isinstance(v, dt.date):
            cell.number_format = DATEFMT
            cell.alignment = CTR_TOP
        elif isinstance(v, (int, float)):
            cell.alignment = CTR_TOP
        else:
            cell.alignment = WRAP_TOP
        if fill:
            cell.fill = fill

def set_widths(ws, c0, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(c0 + i)].width = w

def title_block(ws, c0, ncol, subtitle):
    banner(ws, 1, c0, ncol, "VendorPulse  -  EGB / QBR Process Automation", fill=TITLE_FILL, font=TITLE_FONT, height=30)
    banner(ws, 2, c0, ncol, subtitle, fill=SUB_FILL, font=SUBT_FONT, height=16)
    banner(ws, 3, c0, ncol,
           "Phase 0 pre-mobilisation readiness window from Thu 18-Jun-2026  |  T-0 = Mon 22-Jun-2026  |  4 working weeks (20 days): Wk1 22-26 Jun | Wk2 29 Jun-3 Jul | Wk3 6-10 Jul | Wk4 13-17 Jul (deployment & warranty)  |  CR-9 EGB",
           fill=SUB_FILL, font=Font(name=FONT, size=9, italic=True, color="FF1F4E78"), height=16)

# =================================================================================== #
#  DETAILED PLAN  -  main task sheet                                                  #
# =================================================================================== #
# Columns: Sr | Phase | Workstream | Task / Deliverable | Owner | Start | Days | End |
#          Dependency | Required By / Milestone | Remarks
DCOLS = ["Sr", "Phase", "Workstream", "Task / Deliverable",
         "Start", "Days", "End", "Dependency", "Required By / Milestone", "Remarks"]
DWID  = [4.5, 9, 16, 54, 10.5, 6, 10.5, 30, 22, 56]

# Each task: (workstream, task, owner, start, days, end, dependency, required_by, remarks)
PRE = "Before Day 1"

# ------- Phase 0 : Pre-Mobilisation (Shell-led; runs in parallel with contract negotiation)
P0 = [
 # Identity (Entra ID)
 ("Identity (Entra ID)",
  "Two app registrations (NonProd + Prod) with certificate-based credentials; cert as Key Vault certificate object; 12-month auto-renewal pipeline",
  "Shell Entra ID admin", d(2026,6,8), 3, None,
  "Azure subscription",
  "HARD Day-1 blocker (T-0)",
  "App-only Graph auth + SSO. Cert Shell-generated (preferred) or Zensar-generated, thumbprint returned. Nothing integration-related can start without it."),
 ("Identity (Entra ID)",
  "Admin consent on Graph application permissions: Mail.Send, Calendars.ReadWrite, OnlineMeetings.ReadWrite.All, User.Read.All, MailboxSettings.Read, Group.Read.All",
  "Shell IT Security + Entra admin", d(2026,5,11), 20, None,
  "App registrations; Application Access Policy (coupled)",
  "By end of Wk1 (Fri 26-Jun)",
  "LONG POLE - 2-4 wk security review; the single largest schedule risk. Start ~4-6 wks ahead. In-app form avoids Mail.Read/ReadWrite (narrower consent). May be in progress on Day 1; non-prod uses mocks/dev scopes meanwhile."),
 ("Identity (Entra ID)",
  "OIDC client + redirect URIs configured for SSO (per environment)",
  "Shell Entra admin", d(2026,6,15), 2, None,
  "App registrations",
  "HARD Day-1 blocker (T-0)",
  "Required for any authenticated UI access via Entra SSO."),
 ("Identity (Entra ID)",
  "Entra security groups for the four app roles (VMO coordinator, VMO admin, executive sponsor, viewer)",
  "Shell Entra admin", d(2026,6,15), 2, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "RBAC enforcement; role-gated UI and approval routing."),
 ("Identity (Entra ID)",
  "Exchange Application Access Policy scoping Mail.Send / Calendars to the single service mailbox",
  "Shell Exchange / Messaging admin", d(2026,6,12), 3, None,
  "Service mailbox",
  "HARD Day-1 blocker (T-0)",
  "Least-privilege constraint; couple with admin consent as one security-review package or tenant-wide consent is likely refused."),
 ("Identity (Entra ID)",
  "Conditional Access exemption (or confirmed sign-in path) for the service principal",
  "Shell IT Security", d(2026,6,12), 5, None,
  "App registrations",
  "PROD by Wk4 (non-prod may bypass)",
  "Otherwise CA silently blocks outbound mail/calendar. Non-prod can use a confirmed sign-in path; required for production at the Wk4 security gate."),
 # M365 messaging & calendar
 ("M365 Messaging",
  "Dedicated service mailbox (e.g. vendorpulse-svc@shell.com) with M365 E3/E5 licence",
  "Shell Exchange admin", d(2026,6,12), 5, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "From-identity for all outbound Graph mail; calendar host for Teams meetings. ~$23 (E3) / ~$57 (E5) per user/month."),
 ("M365 Messaging",
  "Mailbox storage allocation (~10 GB) + retention policy (auto-archive sent items after ~1 year)",
  "Shell Exchange admin / Compliance", d(2026,6,12), 3, None,
  "Service mailbox",
  "By Wk1",
  "Keeps the mailbox un-throttled across cycles; supports audit-trail and retention compliance."),
 ("M365 Messaging",
  "Teams meeting policy permitting the service mailbox to create online meetings via Graph",
  "Shell Teams / Exchange admin", d(2026,6,12), 5, None,
  "Service mailbox",
  "By Wk1 (Module A)",
  "Module A and C meeting creation; OnlineMeetings provisioning."),
 ("M365 Messaging",
  "External / guest meeting-policy confirmation (can the mailbox invite vendor external attendees?)",
  "Shell Teams / IT Security", d(2026,6,12), 5, None,
  "Service mailbox",
  "By Wk1 (decision); pilot by Wk4",
  "If blocked, Module A cannot include vendor (external) attendees in the EGB/QBR."),
 ("M365 Messaging",
  "Litmus / Email-on-Acid account for cross-client email-render QA",
  "Shell or Zensar (procured)", d(2026,6,22), 3, None,
  "-",
  "By Day 8 (template QA)",
  "~$99/month. Validates Outlook desktop / OWA / mobile rendering of brand-signed templates."),
 # Azure infrastructure
 ("Azure Platform",
  "Azure subscription with deploy access for Zensar engineers",
  "Shell Azure Cloud admin", d(2026,6,8), 5, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "All deployment, IaC, CI/CD. No environment can be stood up without it."),
 ("Azure Platform",
  "Resource groups (nonprod + prod)",
  "Shell Azure Cloud admin", d(2026,6,15), 1, None,
  "Azure subscription",
  "HARD Day-1 blocker (T-0)",
  "Environment isolation; all resource provisioning."),
 ("Azure Platform",
  "App Service Plan - P1v3 (prod) and B2 (nonprod), Linux containers",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Resource groups",
  "HARD Day-1 blocker (T-0)",
  "Application hosting. ~$220/month (P1v3 prod); B2 nonprod modest."),
 ("Azure Platform",
  "Azure Database for PostgreSQL Flexible Server (HA / zone-redundant for prod)",
  "Shell Azure Cloud admin + DBA", d(2026,6,12), 5, None,
  "Resource groups",
  "HARD Day-1 blocker (T-0)",
  "Target for the Wk1 JSON->Postgres migration; all persistence."),
 ("Azure Platform",
  "Azure Key Vault",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Resource groups",
  "HARD Day-1 blocker (T-0)",
  "Secret storage; Managed-Identity retrieval; certificate object; secret-scrub remediation."),
 ("Azure Platform",
  "Azure Container Registry",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Resource groups",
  "HARD Day-1 blocker (T-0)",
  "Container image build/push for App Service deployment."),
 ("Azure Platform",
  "Azure Front Door + WAF policy",
  "Shell Azure Cloud admin / Networking", d(2026,6,12), 4, None,
  "Resource groups",
  "By Wk1 (cutover in Wk4)",
  "Public ingress, edge protection, TLS termination, cutover slot-swap."),
 ("Azure Platform",
  "App Insights + Log Analytics workspace",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Resource groups",
  "By Wk1 (gate in Wk3)",
  "Observability, alerts, immutable audit trail; feeds the Wk3 observability gate."),
 ("Azure Platform",
  "Private endpoints + VNet integration (Postgres, Key Vault)",
  "Shell Networking", d(2026,6,8), 7, None,
  "Resource groups; Postgres; Key Vault",
  "HARD Day-1 blocker (T-0)",
  "App cannot retrieve secrets via MI or reach the DB at start-up without it. Smoke-test on Day 1."),
 ("Azure Platform",
  "Managed Identity role assignments (Key Vault Secrets User, Postgres grants, Log Analytics)",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Key Vault; Postgres; Log Analytics",
  "HARD Day-1 blocker (T-0)",
  "Secret-less runtime auth; app start-up."),
 ("Azure Platform",
  "Outbound NAT / egress allow-listing to Microsoft Graph + LLM endpoints (Anthropic / Azure OpenAI)",
  "Shell Networking / IT Security", d(2026,6,1), 7, None,
  "VNet",
  "HARD Day-1 blocker (T-0)",
  "HIDDEN LONG POLE (3-7 days). All Graph and LLM calls fail from inside the VNet without it. Validate from non-prod in Wk1."),
 ("Azure Platform",
  "Azure DNS / CNAME + public TLS certificate",
  "Shell Networking / Azure admin", d(2026,6,29), 5, None,
  "Front Door",
  "By Wk4 (DNS/TLS cutover)",
  "Public hostname for the Wk4 cutover - required in Wk4, not Day 1."),
 ("Azure Platform",
  "Azure Policy / tagging compliance",
  "Shell Azure Cloud admin", d(2026,6,15), 2, None,
  "Resource groups",
  "HARD Day-1 blocker (T-0)",
  "Deployments must pass governance or resources are rejected/flagged."),
 # LLM provider
 ("LLM Provider",
  "Anthropic enterprise account + DPA (zero-retention, no-training, EU-residency) + API key in Key Vault + monthly token budget  [RECOMMENDED]",
  "Shell Procurement + Legal/Privacy", d(2026,5,11), 28, None,
  "-",
  "Initiate by T-0; key for Wk2 LLM features",
  "LONG POLE - 4-8 wk contract/DPA. Powers all LLM text (Modules C, D, E, F). ~$1,000/month token budget. Provider-abstracted, so choice is reversible by config."),
 ("LLM Provider",
  "OR Azure OpenAI quota approval + gpt-4o-class chat deployment (NOT preview/computer-use) + key in Key Vault  [config-only alternative]",
  "Shell Azure admin + Procurement", d(2026,5,25), 15, None,
  "-",
  "Initiate by T-0",
  "Same LLM features via the alternative provider behind the single LLMProvider abstraction (switch is config-only). 1-3 wk quota; confirm tool-calling-capable gpt-4o-class in West Europe."),
 # Zensar tooling
 ("Delivery Tooling",
  "Developer laptops for every engineer - full local env (Python 3.11+, Node.js, Docker Desktop, Git) + admin rights",
  "Zensar (Shell device-policy aligned)", d(2026,6,8), 10, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "All local development and build. Without admin rights, Docker/dependency setup is blocked. Zensar-controlled go/no-go gate."),
 ("Delivery Tooling",
  "Approved IDE - VS Code (or JetBrains) with required extensions",
  "Zensar / Shell IT approval", d(2026,6,8), 5, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "Engineer productivity; sanctioned toolchain. Zensar-controlled go/no-go gate."),
 ("Delivery Tooling",
  "Claude Code / Claude for Work licences + Anthropic API developer key (AI-assisted development)",
  "Zensar Procurement / Anthropic", d(2026,6,8), 10, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "The compressed 20-day plan depends on AI-assisted pair-programming velocity. If absent, the plan slips. 1-2 wks if new."),
 ("Delivery Tooling",
  "Source control + CI/CD platform (Azure DevOps or GitHub Enterprise): engineers added as guests, repo created, branch protection, CI/CD service principal",
  "Shell IT / DevOps", d(2026,6,8), 7, None,
  "Azure subscription",
  "HARD Day-1 blocker (T-0)",
  "Day-1 branch & secret-scrub; all commits, pipelines, deployments. Guest onboarding 3-7 days, can be slow."),
 ("Delivery Tooling",
  "CI quality/security gates: SonarQube/SonarCloud, secret scanning (TruffleHog/GitLeaks), container scanning (Trivy/Snyk), Dependabot/Renovate",
  "Zensar / Shell DevOps", d(2026,6,15), 5, None,
  "CI/CD platform",
  "By Wk1 (secret-scrub) / Wk4 (sign-off)",
  "Quality and security gates in CI; secret-scrub verification; supports the Wk4 security sign-off."),
 ("Delivery Tooling",
  "Network access / VPN / device-compliance posture to reach Shell's Azure tenant, repo and non-prod Graph",
  "Shell IT Security / Networking", d(2026,6,1), 10, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "HIDDEN LONG POLE (5-10 days). If unresolved, engineers cannot work against Shell environments at all."),
 ("Delivery Tooling",
  "(Optional) Playwright cloud / BrowserStack for broader cross-browser UI checks",
  "Zensar", d(2026,6,22), 3, None,
  "-",
  "By Wk3 (best-effort)",
  "Best-effort under the compressed timeline; not a blocker."),
 # People & decisions
 ("People & Governance",
  "Shell roles NAMED & available: Executive Sponsor (~2h/wk), Product Owner / VMO lead (daily), IT Architecture & IT Security liaisons, Entra/Exchange/Azure admins, DBA, Networking, Procurement, Legal/Privacy, Brand/Comms",
  "Shell", d(2026,6,15), 5, None,
  "-",
  "HARD Day-1 blocker (PO, IT Arch named)",
  "Any unavailable role can stall a dependency on its critical path. PO/VMO lead and IT Architecture liaison are hard Day-1 named blockers."),
 ("People & Governance",
  "3 named VMO coordinators identified for UAT (active Weeks 3-4)",
  "Shell", d(2026,6,15), 5, None,
  "-",
  "Named by T-0; active Wk3-4",
  "Wk3 UAT, Wk4 training & handover, pilot acceptance."),
 ("People & Governance",
  "Pre-Day-1 decisions LOCKED: LLM provider; hosting region; service-mailbox name & email from-identity / display name; DB choice + CMK yes/no; CI/CD platform; vendor-master format (CSV/Excel); external-attendee policy",
  "Shell + Zensar SA", d(2026,6,15), 5, None,
  "-",
  "HARD Day-1 blocker (T-0)",
  "Platform-level decisions must be settled pre-mob so the Day-2 checkpoint only confirms design-level items. CMK deferred; platform-managed keys for pilot."),
 ("People & Governance",
  "Legal/Privacy: PII / data-residency / audit-retention sign-off path agreed (agent_runs holds vendor names + scorecard comments)",
  "Shell Legal/Privacy + IT Security", d(2026,5,11), 28, None,
  "LLM DPA",
  "Go-live precondition (Wk4)",
  "Until sign-off, auditability of agent_runs vs Shell retention/residency is asserted by design. Start with the DPA, 4-8 wks ahead."),
]

# ------- Phase 1 : Week 1 (Days 1-5) - Foundations, Migration & Design Alignment
P1 = [
 ("DevOps",
  "Branch from POC; scrub POC secrets from Git history; rotate ALL secrets into Key Vault",
  "Zensar DevOps", W1[0], 1, W1[0],
  "CI/CD repo + Key Vault (P0)",
  "Day 1",
  "First task. Verified by secret-scanning in CI. No secrets in code/.env thereafter."),
 ("DevOps",
  "Smoke-test Managed-Identity secret retrieval + private-endpoint connectivity to Postgres & Key Vault from non-prod",
  "Zensar DevOps", W1[0], 1, W1[0],
  "Private endpoints + MI roles (P0)",
  "Day 1",
  "Confirms the secret-less runtime auth path before build starts."),
 ("DevOps",
  "Validate egress allow-list connectivity to Microsoft Graph + LLM endpoints from non-prod",
  "Zensar DevOps", W1[0], 1, W1[0],
  "Egress allow-listing (P0)",
  "Day 1",
  "De-risks the hidden network long pole early in Week 1."),
 ("Design",
  "Day-2 Design Alignment Checkpoint - lock ~12 decisions: scorecard mechanism (native in-app form), 4-category/16-parameter taxonomy (Risk & Compliance, Performance, Commercial, Relationship), brief/minutes output formats, vendor-master source, approval routing",
  "Zensar SA + Shell PO / IT Arch", d(2026,6,23), 1, d(2026,6,23),
  "Pre-Day-1 decisions locked (P0)",
  "GATE - Day 2 (Tue 23-Jun)",
  "Confirms design-level items only; platform-level decisions assumed already locked. Taxonomy supersedes the POC's 5-category interim set."),
 ("Backend",
  "Design PostgreSQL schema (~15 tables, Alembic) and migrate flat JSON/SQLite persistence (asyncpg); data cleansing + defaults for missing fields",
  "Zensar Backend", W1[0], 4, d(2026,6,25),
  "Postgres provisioned (P0); Day-2 taxonomy",
  "Days 1-4",
  "~13 POC entities -> ~15 tables. Additive-only migrations after the Day-13 code freeze."),
 ("Backend",
  "Wire Azure Key Vault with Managed Identity for runtime secret retrieval across all services",
  "Zensar Backend / DevOps", W1[0], 2, d(2026,6,23),
  "Key Vault + MI roles (P0)",
  "Days 1-2",
  "No secrets in configuration files or source control."),
 ("Backend",
  "Remove all Google / Gmail / Google-Forms code entirely from codebase and runtime",
  "Zensar Backend", W1[0], 3, d(2026,6,24),
  "Branch created",
  "Days 1-3",
  "Acceptance criterion: Google stack fully removed."),
 ("Backend",
  "Refactor GraphService to app-only certificate authentication (MSAL client-credentials, /users/{id} endpoints, retry/backoff on 429/5xx)",
  "Zensar Backend", d(2026,6,23), 4, d(2026,6,26),
  "App reg + cert (P0); admin consent for live test",
  "Days 2-5",
  "Replaces the POC's pasted delegated bearer token. Dev against mocks/dev scopes until consent clears by Friday."),
 ("Backend / Auth",
  "Implement Entra ID OIDC Single Sign-On + RBAC (VMO coordinator, VMO admin, executive sponsor, viewer); server-side enforcement",
  "Zensar Backend + Frontend", d(2026,6,23), 4, d(2026,6,26),
  "OIDC client + security groups (P0)",
  "Days 2-5",
  "Roles carried in the token; UI reflects role, never decides it."),
 ("AI Platform",
  "Implement LLMProvider abstraction; add recommended Anthropic Claude provider (Azure OpenAI gpt-4o-class config-only alternative). Self-implemented tool-calling agent layer - MAF SDK NOT used. LLM features disable-by-configuration (feature flag); graceful degradation if provider unavailable",
  "Zensar Backend", d(2026,6,24), 2, d(2026,6,26),
  "LLM key in Key Vault (P0); egress",
  "Days 3-5",
  "Switching providers is config-only. Agents use the model's tool-calling for schema-validated structured output. Graceful degradation: the deterministic workflow proceeds if the LLM is down - AI text is optional, never blocking; coordinators can proceed manually."),
 ("Module A",
  "Module A (Scheduling & Coordination) end-to-end on REAL Graph: attendee/roster refresh, findMeetingTimes against real Shell mailboxes, deterministic slot ranking (organiser + sponsor free = hard constraint; soft scoring on attendance %, calendar conflicts, timezone fit), Teams meeting + invite, RSVP tracking",
  "Zensar Backend + Frontend", d(2026,6,24), 3, W1[1],
  "Graph admin consent clears (P0); service mailbox; Teams policy",
  "By Fri Day 5 (milestone)",
  "Only AI contribution is optional outreach-text polish. Invite sent only after coordinator approval."),
 ("Gate",
  "GATE - Design freeze (end of Week 1). Later changes route to the defect or Phase-2 path",
  "Zensar SA + Shell IT Arch", W1[1], 1, W1[1],
  "Module A demo on real Graph; Day-2 decisions",
  "GATE - Day 5 (Fri 26-Jun)",
  "Module A demonstrated end-to-end on real Graph by Friday; scope frozen."),
]

# ------- Phase 2 : Week 2 (Days 6-10) - Functional Completion
P2 = [
 ("Module B",
  "Module B (Scorecard Input & Validation): native in-app form (1-5 across 4 categories / 16 parameters: Risk & Compliance, Performance, Commercial, Relationship) persisted in Postgres; Graph-mail requests + escalating reminders; deterministic validation (range, mandatory comment on 1/5, ~1.5-sigma outlier flag); compile internal-vs-vendor scorecard",
  "Zensar Backend + Frontend", W2[0], 4, d(2026,7,2),
  "Taxonomy locked (Day 2); Graph mail; Postgres",
  "Days 6-10",
  "Replaces the Google Form + polling (supersedes the POC's 5-category interim set). LLM performs no part of scoring/validation. Scorecard request & reminder emails dispatched only after coordinator approval (approval gate)."),
 ("Module C",
  "Module C (Internal Alignment): deterministic score deltas vs previous cycle; divergence flag where spread >= 1.5; LLM 'what changed' narrative + action-item extraction from notes; schedule internal-only Teams meeting via Graph",
  "Zensar Backend", W2[0], 4, d(2026,7,2),
  "LLMProvider (Wk1); Graph; Postgres",
  "Days 6-10",
  "LLM outputs reviewed/edited by coordinator; alignment invite sent only after approval."),
 ("Module D",
  "Module D (Vendor Prep): AI vendor brief (overall score/trend, per-category ratings, concerns, positives, open actions); 3 response stances (Factual / Neutral / Escalation) per pushback; deterministic guardrail suppresses AI drafting for legal-review items; unresolved-item tracking",
  "Zensar Backend + Frontend", W2[0], 4, d(2026,7,2),
  "LLMProvider (Wk1); Postgres",
  "Days 6-10",
  "Internal-facing advisory material; no outbound comms of its own."),
 ("Module E",
  "Module E (EGB/QBR Meeting): pre-meeting facilitator briefing; live notes (5 types: Question/Objection/Decision/Appreciation/Action); pasted-transcript parse into structured notes; formal minutes; action-item extraction; distribute approved minutes via Graph mail",
  "Zensar Backend + Frontend", W2[0], 4, d(2026,7,2),
  "LLMProvider; Graph mail",
  "Days 6-10",
  "Det-vs-AI split: LLM parses the transcript and drafts the facilitator briefing + formal minutes; the 5-type note classification, action-item register and workflow-state advance stay deterministic. Teams transcript ingestion out of scope - transcripts pasted in. Minutes held behind the approval gate."),
 ("Module F",
  "Module F (Cross-Cycle Memory & Trends): deterministic multi-cycle trends, recurring-issue detection (category below threshold for >= 2 consecutive cycles), trajectory classification (improving/stable/declining), LLM executive leadership brief, cross-vendor comparison",
  "Zensar Backend", W2[0], 4, d(2026,7,2),
  "Postgres history; LLMProvider",
  "Days 6-10",
  "Read-only/analytical; history now in Postgres rather than flat JSON. Only the leadership brief is LLM-generated; any leadership brief shared externally still follows the standard approval gate."),
 ("Frontend",
  "Shell-branded UI applied across the React 19 SPA (Design System + ApprovalPanel)",
  "Zensar Frontend", W2[0], 3, d(2026,7,1),
  "Brand/Comms assets (P0)",
  "Days 6-8",
  "ApprovalPanel is the primary HITL surface for every AI draft."),
 ("Frontend",
  "Outlook-friendly HTML email templates (table-based layout, inline styles, plain-text fallback); cross-client render QA (Litmus / Email-on-Acid)",
  "Zensar Frontend + QA", W2[0], 4, d(2026,7,2),
  "Brand sign-off by Day 8; Litmus acct (P0)",
  "Day 8 (brand sign-off); QA by Day 10",
  "Verified render on Outlook desktop / OWA / mobile."),
 ("Admin",
  "Admin module: vendor master, user-role view, LLM budget panel, audit-log viewer, system-health view",
  "Zensar Backend + Frontend", W2[0], 4, d(2026,7,2),
  "Postgres; audit model",
  "Days 6-10",
  "Operability surfaces for Shell IT Ops."),
 ("Backend",
  "Audit model: agent_runs, external_calls, security_events; append-only design mirrored to Log Analytics; correlation IDs propagated end-to-end; approver identity + timestamp recorded per outbound action (human accountability)",
  "Zensar Backend", W2[0], 4, d(2026,7,2),
  "App Insights + Log Analytics (P0)",
  "Days 6-10",
  "Append-only/auditability validated in Wk3. Each outbound communication is attributable to the named approving coordinator and timestamped. agent_runs holds vendor names + comments (Legal/Privacy sign-off)."),
 ("QA / Frontend",
  "Automated accessibility checks (axe / Lighthouse) on the React 19 SPA",
  "Zensar QA + Frontend", d(2026,6,30), 4, d(2026,7,3),
  "Branded UI (Day 8)",
  "Days 7-10 (automated only)",
  "Automated checks at launch; a formal manual WCAG audit is deferred to the warranty period / Phase 2 (Out of Scope at launch)."),
 ("Hardening",
  "Hardening: rate limiting, token-budget guards + per-call caps, retry/backoff, correlation IDs",
  "Zensar Backend / Tech Lead", d(2026,7,1), 3, d(2026,7,3),
  "All six modules",
  "Days 8-10",
  "Bounds LLM spend and protects against Graph throttling."),
 ("QA",
  "Add targeted automated test coverage (deterministic logic + smoke tests) - begins Wk2, continues Wk3",
  "Zensar QA", d(2026,6,30), 3, d(2026,7,3),
  "Modules in build",
  "Days 7-10",
  "POC has no test suite; focused coverage compensates under the compressed timeline."),
 ("Milestone",
  "Milestone - all six agents (A-F) functional end-to-end; functional build substantially complete",
  "Zensar Tech Lead + Delivery Mgr", W2[1], 1, W2[1],
  "Modules A-F; UI; admin; audit",
  "End of Wk2 (Fri 3-Jul)",
  "Not the code freeze. Remaining completion + final hardening continue into early Wk3, up to the Day-13 code freeze."),
]

# ------- Phase 3 : Week 3 (Days 11-15) - Completion, mid-week Code Freeze, Stabilisation & UAT
P3 = [
 ("Completion",
  "Final functional completion, integration polish & hardening (rate limiting, token-budget guards, retry/backoff, correlation IDs) - close out any remaining build",
  "Zensar Backend + Frontend / Tech Lead", W3[0], 3, d(2026,7,8),
  "Functional build (Wk2)",
  "Days 11-13 (by code freeze)",
  "Any build work not finished in Wk2 is completed here, ahead of the mid-week code freeze."),
 ("UAT",
  "User Acceptance Testing with the 3 named VMO coordinators (all six modules A-F)",
  "Zensar QA + Shell VMO coordinators", W3[0], 5, W3[1],
  "Modules functional; coordinators available (P0)",
  "Days 11-15",
  "Risk-based manual UAT; begins on completed modules and intensifies after the Day-13 code freeze (defect fixes only thereafter). No production go-live this week."),
 ("UAT",
  "Defect triage across P1/P2/P3 with P1 resolved same-day; daily defect burn-down",
  "Zensar QA + Tech Lead", W3[0], 5, W3[1],
  "UAT in progress",
  "Days 11-15",
  "Daily stand-up reprioritisation; CAB contingency slot held for a late major defect."),
 ("Ops",
  "Complete observability, alerting and the operational runbook v1.0",
  "Zensar DevOps + Tech Lead", W3[0], 5, W3[1],
  "App Insights (P0)",
  "Days 11-15",
  "Feeds the Wk3 observability gate and the Wk4 handover."),
 ("Gate",
  "GATE - Code freeze (middle of Week 3, Wed 8-Jul, Day 13). No further feature development; stabilisation & defect-fixes only",
  "Zensar Tech Lead + Delivery Mgr", d(2026,7,8), 1, d(2026,7,8),
  "All six agents functional; final hardening complete",
  "GATE - Day 13 (Wed 8-Jul)",
  "From this point: defects only. Schema effectively frozen (additive-only migrations)."),
 ("QA",
  "Regression + smoke testing of deterministic logic (slot ranking, score validation, workflow transitions)",
  "Zensar QA", d(2026,7,8), 3, W3[1],
  "Code freeze (Day 13)",
  "Days 13-15 (post-freeze)",
  "Validates the deterministic-first core behaves identically after re-platforming."),
 ("QA / Security",
  "Validate platform-layer capabilities: audit append-only behaviour, RBAC, app-only Graph auth, email rendering",
  "Zensar QA + SA", d(2026,7,8), 3, W3[1],
  "Audit model; RBAC; Graph auth",
  "Days 13-15 (post-freeze)",
  "These platform traits are asserted-by-design in the POC and validated here before cutover."),
 ("Gate",
  "GATE - UAT sign-off (end of Week 3). All P1/P2 UAT defects closed; release candidate approved for deployment",
  "Shell PO / IT Security + Zensar", W3[1], 1, W3[1],
  "UAT complete; defects closed",
  "GATE - Day 15 (Fri 10-Jul)",
  "Exit with a tested release candidate. P3 defects triaged into warranty/Phase 2."),
]

# ------- Phase 4 : Week 4 (Days 16-20) - Security Sign-off, Deployment & Defect-Warranty
P4 = [
 ("Security",
  "Shell IT Security review + sign-off on Graph permission scopes, Application Access Policy, RBAC, audit posture, and single-tenant isolation / data-residency (no Shell data outside Shell's boundary)",
  "Shell IT Security + Zensar SA", W4[0], 2, d(2026,7,14),
  "UAT sign-off; security artefacts",
  "GATE - early Wk4 (Mon-Tue)",
  "Engage Security liaison from Day 1 to avoid a late objection blocking cutover. Reviews deployment topology + data-flow for tenant isolation."),
 ("Deployment Prereqs",
  "Confirm production prerequisites landed: Conditional Access exemption for PROD service principal; production DNS hostname + public TLS certificate; Front Door/WAF prod config",
  "Shell Networking / Entra + Zensar DevOps", W4[0], 2, d(2026,7,14),
  "CA path, DNS/TLS, Front Door (P0)",
  "Before cutover (Wk4)",
  "These were 'needed by Wk4' pre-mob items; confirm complete before the cutover window."),
 ("Pilot Data",
  "Vendor-master import (CSV/Excel) for the agreed single pilot vendor",
  "Shell PO + Zensar Backend", W4[0], 2, d(2026,7,14),
  "Shell provides pilot vendor data (P0)",
  "Before pilot kick-off (Wk4)",
  "Late/incomplete vendor data delays the Wk4 pilot kick-off and handover even if the platform is live."),
 ("Gate",
  "GATE - CAB (Change Advisory Board) approval of the production change",
  "Shell CAB + Exec Sponsor", d(2026,7,14), 1, d(2026,7,15),
  "Security sign-off",
  "GATE - Wk4 (Wed)",
  "Change Advisory Board approval is a precondition for cutover."),
 ("Deployment",
  "Production cutover / go-live: DNS change, deployment slot-swap, 60-minute observation window; tested rollback via slot-swap",
  "Zensar DevOps + Shell Networking", d(2026,7,15), 1, d(2026,7,15),
  "CAB approval; DNS/TLS; Front Door",
  "GATE - Wk4 (Wed-Thu)",
  "System live in Shell's Azure subscription on App Service (Linux containers) behind Front Door + WAF."),
 ("Pilot",
  "Pilot vendor cycle kick-off in production, pair-shadowed by the Zensar Tech Lead",
  "Zensar Tech Lead + Shell VMO coordinators", d(2026,7,16), 2, W4[1],
  "Cutover complete; vendor-master import",
  "Wk4 (Thu-Fri)",
  "One real pre-agreed pilot vendor cycle progressing with no open P1 issues."),
 ("Handover",
  "Coordinator training (recorded) + handover: runbook v1.0, documentation and the design decision log committed to Shell's repository",
  "Zensar Tech Lead + SA", d(2026,7,16), 2, W4[1],
  "Cutover; runbook v1.0 (Wk3)",
  "GATE - Wk4 (Thu-Fri)",
  "Training delivered to the 3 named VMO coordinators; docs handed to Shell IT Ops."),
 ("Warranty",
  "Defect-warranty support begins (4-week period): Zensar fixes P1/P2 defects (P1 same-day where feasible) while Shell IT Ops runs BAU; P3 to backlog/Phase 2",
  "Zensar (warranty) + Shell IT Ops", W4[1], 1, W4[1],
  "Go-live",
  "Begins at go-live (Wk4)",
  "Warranty start/end dates agreed in writing at handover. Post-freeze changes route to defect path or Phase 2."),
]

PHASES_DETAILED = [
 ("PHASE 0  -  Pre-Mobilisation  (before T-0; Shell-led, runs in parallel with contract negotiation. Long-lead items start ~4-6 weeks ahead)", P0),
 ("PHASE 1  -  Week 1 (Days 1-5, 22-26 Jun)  -  Foundations, Migration & Design Alignment  [Development]", P1),
 ("PHASE 2  -  Week 2 (Days 6-10, 29 Jun-3 Jul)  -  Functional Build (all six agents)  [Development]", P2),
 ("PHASE 3  -  Week 3 (Days 11-15, 6-10 Jul)  -  Completion, mid-week Code Freeze (Wed 8-Jul), Stabilisation & UAT  [Development - no go-live]", P3),
 ("PHASE 4  -  Week 4 (Days 16-20, 13-17 Jul)  -  Security Sign-off, Deployment & Defect-Warranty", P4),
]

def fill_for(required_by, workstream):
    if workstream == "Gate":
        return GATE_FILL
    if isinstance(required_by, str) and "HARD Day-1" in required_by:
        return BLOCK_FILL
    return None

def build_detailed(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Implementation Plan"
    ws.sheet_view.showGridLines = False
    c0 = 2
    ncol = len(DCOLS)
    set_widths(ws, c0, DWID)
    title_block(ws, c0, ncol, "DETAILED Implementation Plan  -  full task breakdown, dependencies, owners & dated milestones")
    info = Font(name=FONT, size=9, color="FF1F4E78")
    # legend
    banner(ws, 4, c0, ncol,
           "Legend:  orange = HARD Day-1 blocker (must be complete by T-0)   |   gold = quality GATE   |   Phase 0 is a parallel readiness window dated 18-19 Jun - external lead times are in the Pre-Mob Checklist 'Lead time' column; the two long poles (Graph admin consent, LLM contract/DPA) must be INITIATED ~4-6 wks before T-0.   Agent layer = self-implemented tool-calling (no MAF SDK); LLM provider-abstracted (Anthropic Claude recommended, Azure OpenAI config-only).",
           fill=SUB_FILL, font=info, height=28)
    banner(ws, 5, c0, ncol,
           "Team (Zensar, 100% allocated, no planned leave Wk1-3; deployment & warranty Wk4):  1 Delivery Mgr, 1 Solution Architect, 1 Tech Lead, 2 Backend, 2 Frontend, 1 QA, 1 DevOps/Cloud.  Shell supporting: Exec Sponsor (~2h/wk), Product Owner/VMO lead (daily), IT Architecture & Security liaisons, Entra/Exchange/Azure admins, DBA, Networking, Procurement, Legal/Privacy, 3 VMO coordinators, Brand/Comms.",
           fill=SUB_FILL, font=info, height=28)
    banner(ws, 6, c0, ncol,
           "Scope basis:  POC functional behaviour (Modules A-F) is accepted as the baseline; effort sizes the integration & platform re-engineering only - NOT rework of functional logic, the 12-state workflow, the human-approval gate or the deterministic-first rules.  New/changed functional requirements are out of scope and require re-estimation.  Launch is English-only, web-only (desktop browser) and single-tenant (Shell's Azure subscription + M365 tenant only).",
           fill=SUB_FILL, font=info, height=28)
    banner(ws, 7, c0, ncol,
           "Architectural invariants (unchanged across A-F):  (1) 12-state forward-only workflow - CYCLE_CREATED -> ATTENDEE_REFRESH_SENT -> AVAILABILITY_COLLECTED -> MEETING_SCHEDULED -> SCORECARD_REQUEST_SENT -> SCORECARD_COLLECTION -> SCORECARD_COMPILED -> INTERNAL_ALIGNMENT -> VENDOR_PREP -> MEETING_IN_PROGRESS -> POST_MEETING_COMPLETE -> ARCHIVED (invalid transitions rejected);  (2) human-approval gate on every outbound action;  (3) deterministic-first, AI-second.",
           fill=SUB_FILL, font=info, height=28)
    hdr = 9
    header(ws, hdr, c0, DCOLS)
    r = hdr + 1
    sr = 1
    PREMOB_START = d(2026, 6, 18)   # pre-mobilisation readiness window opens (Thu before T-0)
    PREMOB_END   = d(2026, 6, 19)   # ready by Fri 19-Jun; T-0 = Mon 22-Jun
    for phase_title, tasks in PHASES_DETAILED:
        banner(ws, r, c0, ncol, phase_title); r += 1
        is_p0 = phase_title.startswith("PHASE 0")
        for (wsname, task, owner, start, days, end, dep, reqby, remarks) in tasks:
            if is_p0:
                # Phase 0 is a parallel readiness window anchored at 18-Jun; external lead
                # times live in the Pre-Mob Checklist 'Lead time' column and the remarks.
                start, days, end = PREMOB_START, 2, PREMOB_END
            elif end is None and isinstance(start, dt.date):
                end = start + dt.timedelta(days=days)   # indicative finish
            fill = fill_for(reqby, wsname)
            phase_short = phase_title.split("  -  ")[0].replace("PHASE ", "P")
            vals = [sr, phase_short, wsname, task, start, days, end, dep, reqby, remarks]
            write_data_row(ws, r, c0, vals, fill=fill)
            # gate rows: bold the task
            if wsname == "Gate":
                ws.cell(row=r, column=c0+3).font = BOLD_FONT
                for i in range(ncol):
                    ws.cell(row=r, column=c0+i).fill = GATE_FILL
            ws.row_dimensions[r].height = 42
            r += 1; sr += 1
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)
    # ---- supporting sheets
    build_dependency_sheet(wb)
    build_gates_sheet(wb)
    build_risk_sheet(wb)
    build_acceptance_sheet(wb)
    wb.save(path)
    return r - (hdr + 1)

# --------------------------------------------------------------------------- Pre-Mob checklist sheet
def build_dependency_sheet(wb):
    ws = wb.create_sheet("Pre-Mob Checklist")
    ws.sheet_view.showGridLines = False
    cols = ["#", "Category", "Dependency item", "Owner", "Lead time", "Indicative cost", "Blocks / required for", "Day-1 status"]
    wid  = [4, 16, 44, 26, 13, 18, 40, 26]
    c0 = 2
    set_widths(ws, c0, wid)
    title_block(ws, c0, len(cols), "Pre-Mobilisation Dependency Checklist  -  what blocks Day 1 vs what is needed later")
    banner(ws, 4, c0, len(cols),
           "HARD Day-1 blockers must be complete by Monday of Week 1 (T-0 = 22-Jun).  'In progress / by <date>' items may still be open on Day 1 provided they land by the stated milestone.",
           fill=SUB_FILL, font=Font(name=FONT, size=9, color="FF1F4E78"), height=16)
    hdr = 6; header(ws, hdr, c0, cols); r = hdr + 1
    rows = [
      ("Hard Day-1 blockers (complete by Mon of Week 1)", None),
      ("Identity", "Both Entra app registrations (NonProd + Prod) with certificate credentials", "Shell Entra admin", "2-3 days", "No incremental", "App-only Graph auth; all email/calendar/Teams; SSO", "HARD Day-1"),
      ("Identity", "OIDC client + redirect URIs configured for SSO", "Shell Entra admin", "1-2 days", "No incremental", "Any authenticated UI access", "HARD Day-1"),
      ("Identity", "Entra security groups for the four app roles", "Shell Entra admin", "1-2 days", "No incremental", "RBAC; role-gated UI & approval routing", "HARD Day-1"),
      ("M365", "Service mailbox provisioned + Exchange Application Access Policy applied", "Shell Exchange admin", "2-5 days", "~$23-57/user/mo", "From-identity for outbound mail; least-privilege scope", "HARD Day-1"),
      ("Azure", "Subscription + resource groups (nonprod + prod) with Zensar deploy access", "Shell Azure admin", "2-5 days", "Consumption", "All deployment, IaC, CI/CD", "HARD Day-1"),
      ("Azure", "Azure Key Vault provisioned", "Shell Azure admin", "1-2 days", "Negligible", "Secret storage; MI retrieval; cert object", "HARD Day-1"),
      ("Azure", "Azure Database for PostgreSQL Flexible Server provisioned", "Shell Azure admin + DBA", "2-5 days", "In prod estimate", "Wk1 JSON->Postgres migration; all persistence", "HARD Day-1"),
      ("Azure", "App Service Plan (P1v3 prod / B2 nonprod) + Azure Container Registry", "Shell Azure admin", "1-2 days", "~$220/mo prod", "Application hosting; image build/push", "HARD Day-1"),
      ("Azure", "Private endpoints + VNet integration + Managed Identity role assignments", "Shell Networking / Azure admin", "3-7 days", "Modest", "Secret-less runtime auth; DB & KV connectivity", "HARD Day-1"),
      ("Azure", "Outbound NAT / egress allow-listing to Graph + LLM endpoints", "Shell Networking / IT Sec", "3-7 days", "No incremental", "All Graph & LLM calls from the VNet (hidden long pole)", "HARD Day-1"),
      ("Azure", "Azure Policy / tagging compliance", "Shell Azure admin", "1-2 days", "No incremental", "Deployments pass governance", "HARD Day-1"),
      ("LLM", "LLM provider chosen + procurement initiated (contract/DPA in flight)", "Shell Procurement + Legal", "4-8 wks (in flight)", "~$1,000/mo budget", "All LLM text (C, D, E, F); key needed Wk2", "HARD Day-1 (decision)"),
      ("Tooling", "Developer laptops (Python 3.11+, Node, Docker, Git) + admin rights", "Zensar", "Before Day 1", "Zensar-borne", "All local development and build", "HARD Day-1"),
      ("Tooling", "Approved IDE (VS Code / JetBrains) + extensions", "Zensar / Shell IT", "Before Day 1", "Free / licence", "Sanctioned toolchain", "HARD Day-1"),
      ("Tooling", "Claude Code / Claude for Work licences + Anthropic API dev key", "Zensar / Anthropic", "1-2 wks if new", "Per-seat + usage", "AI-assisted pair-programming velocity", "HARD Day-1"),
      ("Tooling", "CI/CD repo created; Zensar engineers as guests; branch protection; CI/CD SP", "Shell IT / DevOps", "3-7 days", "Per licence", "Day-1 branch & secret-scrub; all pipelines", "HARD Day-1"),
      ("Connectivity", "Network / VPN / device-compliance posture to reach tenant, repo, non-prod Graph", "Shell IT Sec / Networking", "5-10 days", "No incremental", "Day-1 connectivity (hidden long pole)", "HARD Day-1"),
      ("People", "Product Owner / VMO lead + IT Architecture liaison named & available", "Shell", "Named pre-Day-1", "-", "Daily decisions; Day-2 checkpoint; design-freeze sign-off", "HARD Day-1"),
      ("Decisions", "Pre-Day-1 decisions locked (LLM provider, region, mailbox name + from-identity/display name, DB+CMK, CI/CD, vendor-master format, external-attendee policy)", "Shell + Zensar SA", "Pre-Day-1", "-", "Keeps the Day-2 checkpoint to design-level items only", "HARD Day-1"),
      ("May be in progress on Day 1, required by the date shown", None),
      ("Identity", "Graph application-permission admin consent (LONG POLE, 2-4 wk review)", "Shell IT Sec + Entra", "2-4 wks", "No incremental", "Module A end-to-end on real Graph", "By Day 5 (Fri 26-Jun)"),
      ("Identity", "Conditional Access exemption / confirmed sign-in path for the SP", "Shell IT Security", "2-5 days", "No incremental", "Service-principal Graph calls in PROD", "PROD by Wk4"),
      ("Azure", "Production DNS hostname + public TLS certificate", "Shell Networking", "2-5 days", "Cert cost", "Wk4 DNS/TLS cutover", "By Wk4"),
      ("Azure", "Azure Front Door + WAF policy + App Insights/Log Analytics", "Shell Azure / Networking", "2-4 days", "In prod estimate", "Public ingress; Wk4 cutover; Wk3 observability gate", "Wk1 build / Wk4 cutover"),
      ("Pilot", "Vendor-master import (CSV/Excel) for the pilot vendor", "Shell PO", "-", "-", "Wk4 pilot vendor cycle kick-off", "By Wk4 kick-off"),
      ("Brand", "Brand / Comms sign-off on Shell-branded email templates", "Shell Brand/Comms", "-", "-", "Templates built & rendered in Wk2", "By Day 8 (Wk2)"),
      ("M365", "Litmus / Email-on-Acid account for cross-client render QA", "Shell or Zensar", "1-3 days", "~$99/mo", "Outlook render acceptance (Day 8 templates)", "By Day 8"),
      ("People", "3 named VMO coordinators for UAT", "Shell", "Named pre-Day-1", "-", "Wk3 UAT; Wk4 training & handover", "Active Wk3-4"),
    ]
    for row in rows:
        if row[1] is None:
            banner(ws, r, c0, len(cols), row[0],
                   fill=(BLOCK_FILL if "Hard" in row[0] else LATER_FILL),
                   font=Font(name=FONT, bold=True, size=10), height=18)
            r += 1; continue
        cat, item, owner, lead, cost, blocks, status = row
        n = r - hdr  # running number-ish
        vals = [n, cat, item, owner, lead, cost, blocks, status]
        fill = BLOCK_FILL if "HARD" in status else LATER_FILL
        write_data_row(ws, r, c0, vals, fill=fill)
        ws.row_dimensions[r].height = 30
        r += 1
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)

# --------------------------------------------------------------------------- gates sheet
def build_gates_sheet(wb):
    ws = wb.create_sheet("Quality Gates")
    ws.sheet_view.showGridLines = False
    cols = ["#", "Gate", "Timing", "Pass condition"]
    wid  = [4, 34, 26, 90]
    c0 = 2; set_widths(ws, c0, wid)
    title_block(ws, c0, len(cols), "Quality Gates  -  five sequential gates; each must pass before the next stage proceeds")
    hdr = 5; header(ws, hdr, c0, cols); r = hdr + 1
    gates = [
      ("Day-2 Design Alignment Checkpoint", "Week 1, Day 2 (Tue 23-Jun)", "~12 design decisions locked (scorecard mechanism, 4-cat/16-param taxonomy, brief/minutes output formats, vendor-master source, approval routing)."),
      ("Design freeze", "End of Week 1 (Fri 26-Jun, Day 5)", "Scope frozen; Module A demonstrated end-to-end on real Graph by Friday. Later changes route to the defect or Phase-2 path."),
      ("Code freeze", "Middle of Week 3 (Wed 8-Jul, Day 13)", "All six agents functional; scorecard collection in chosen mode; Shell-branded UI; Outlook templates; admin module; final hardening complete. From here: defects only - stabilisation & UAT."),
      ("UAT sign-off", "End of Week 3 (Fri 10-Jul, Day 15)", "UAT complete with all P1/P2 defects closed; release candidate approved for deployment. No production go-live this week."),
      ("Security sign-off", "Week 4 (Mon-Tue 13-14 Jul)", "Shell IT Security review passed on Graph scopes, Application Access Policy, RBAC, audit posture; observability, alerts and runbook in place."),
      ("CAB approval & production cutover", "Week 4 (Wed-Thu 15-16 Jul)", "Change Advisory Board approval; production cutover (DNS, slot-swap, 60-min observation); pilot vendor cycle kicked off."),
      ("Handover", "Week 4 (Thu-Fri 16-17 Jul)", "Coordinator training delivered & recorded; runbook v1.0, documentation and decision log handed over; defect-warranty begins."),
    ]
    for i, (g, t, c) in enumerate(gates, 1):
        write_data_row(ws, r, c0, [i, g, t, c], fill=GATE_FILL)
        ws.cell(row=r, column=c0+1).font = BOLD_FONT
        ws.row_dimensions[r].height = 40
        r += 1
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)

# --------------------------------------------------------------------------- risk register sheet
def build_risk_sheet(wb):
    ws = wb.create_sheet("Risk Register")
    ws.sheet_view.showGridLines = False
    cols = ["#", "Category", "Risk", "Likelihood", "Impact", "Mitigation", "Owner"]
    wid  = [4, 22, 46, 12, 10, 56, 26]
    c0 = 2; set_widths(ws, c0, wid)
    title_block(ws, c0, len(cols), "Consolidated Risk Register  -  residual risk concentrated in the pre-mobilisation window")
    hdr = 5; header(ws, hdr, c0, cols); r = hdr + 1
    risks = [
      ("Tooling/velocity", "Developer laptops not provided with full local env + admin rights", "Low", "High", "Confirm laptop/device readiness as a Zensar-controlled pre-mob go/no-go gate; Zensar-managed devices as fallback under Shell device-compliance", "Zensar Delivery Mgr / Shell IT"),
      ("Tooling/velocity", "VS Code / approved IDE + extensions not approved", "Low", "Medium", "Lock IDE choice and extension list with Shell IT before Day 1; go/no-go gate", "Shell IT / Zensar Delivery Mgr"),
      ("Tooling/velocity", "Claude Code / Anthropic developer licences not procured", "Low", "High", "Procure Claude for Work licences + API dev key in pre-mob; go/no-go gate", "Zensar Delivery Mgr / Procurement"),
      ("Tooling/velocity", "Network / VPN / device compliance prevents reaching Shell tenant, repo, non-prod Graph", "Medium", "High", "Agree access posture and VPN/BYOD path in pre-mobilisation", "Shell Networking / Security"),
      ("Tooling/velocity", "Source-control + CI/CD guest onboarding slow (guests, branch protection, CI/CD SP)", "Medium", "High", "Provision repo guest access, branch protection and CI/CD SP pre-Day-1; treat as Day-1 blocker (3-7 day onboarding)", "Shell IT / Zensar DevOps"),
      ("Identity/Graph", "Graph app-only admin consent not granted in time", "High", "High", "Start admin consent ~4-6 wks ahead (2-4 wk review, LONG POLE); track daily; mock only for internal demo", "Shell Entra admin / IT Security"),
      ("Identity/Graph", "Application Access Policy not offered/accepted, so tenant-wide consent is refused", "Medium", "High", "Propose mailbox-scoped Application Access Policy up front; couple policy + consent as one security-review package", "Shell IT Security / Exchange admin"),
      ("Identity/Graph", "Service mailbox / Application Access Policy not provisioned or mis-scoped", "Medium", "High", "Provision vendorpulse-svc mailbox + E3/E5 licence and apply the scoping policy in pre-mob; offer policy up front", "Shell Exchange admin"),
      ("Identity/Graph", "Graph certificate not generated/uploaded in time, or client-secret used against preference", "Medium", "High", "Agree cert ownership (Shell-generated preferred) + thumbprint handover; store as Key Vault cert object; 12-month auto-renewal", "Shell Entra admin / Zensar DevOps"),
      ("Identity/Graph", "Microsoft Graph API throttling (429 / 5xx)", "Medium", "Medium", "Retry/backoff engineered in GraphService; correlation IDs; batch requests where possible", "Zensar Tech Lead"),
      ("Identity/Graph", "Conditional Access blocks the service principal", "Medium", "High", "CA exemption or confirmed sign-in path agreed with Entra admin pre-Day-1", "Shell Entra admin"),
      ("Identity/Graph", "External / guest meeting policy blocks vendor invites", "Medium", "High", "Confirm external-attendee policy for the service mailbox in pre-mobilisation", "Shell Exchange/Teams admin"),
      ("Azure/Network", "Outbound egress / NAT allow-listing to Graph + LLM endpoints not configured", "Medium", "High", "Agree the egress allow-list (Graph + Anthropic/Azure OpenAI) with Shell Networking pre-mob; validate from non-prod in Wk1", "Shell Networking / Zensar DevOps"),
      ("Azure/Network", "Private endpoints / VNet integration to Postgres & Key Vault not in place", "Medium", "High", "Provision private endpoints + VNet integration + MI role assignments pre-mob; smoke-test secret retrieval & DB connectivity Day 1", "Shell Networking / Azure admin / Zensar DevOps"),
      ("Azure/Network", "No SQL schema yet + messy POC SQLite/JSON data (13 entities -> ~15 tables)", "Medium", "Medium", "Alembic schema + asyncpg planned Day 1; data cleansing scoped; defaults for missing fields; additive-only migrations after Day-13 freeze", "Zensar Backend / Tech Lead"),
      ("Azure/Network", "Azure estate incomplete (RGs, Postgres, KV, App Service, Front Door/WAF, App Insights, role assignments, egress)", "Medium", "High", "Complete Azure provisioning in pre-mob; validate MI role assignments Day 1", "Shell Azure admin / Zensar DevOps"),
      ("Azure/Network", "DNS / TLS cutover failure in Week 4", "Low", "High", "Provision DNS+TLS in pre-mob; slot-swap with rollback; 60-min observation window", "Shell Networking / Zensar DevOps"),
      ("LLM/Compliance", "LLM enterprise contract / DPA / EU-residency not signed", "Medium", "High", "Start contract 4-8 wks ahead; LLMProvider abstraction enables config-only switch to Azure OpenAI", "Shell Procurement / Legal/Privacy"),
      ("LLM/Compliance", "Azure OpenAI quota / region approval not granted (if chosen)", "Low-Medium", "Medium", "Submit Microsoft quota request in chosen region 2-4 wks ahead; confirm tool-calling-capable gpt-4o-class in West Europe", "Shell Azure admin / Zensar Tech Lead"),
      ("LLM/Compliance", "Azure OpenAI deployment misconfigured (preview / non-chat model)", "Low", "Medium", "Specify a gpt-4o-class chat deployment explicitly; validate connectivity in Wk1", "Shell Azure admin / Zensar Tech Lead"),
      ("LLM/Compliance", "LLM token-spend overrun", "Medium", "Medium", "Token-budget guards + per-call caps; admin LLM budget panel; monthly budget agreed", "Zensar Tech Lead / Shell PO"),
      ("LLM/Compliance", "LLM provider outage mid-cycle blocks the workflow", "Low", "Medium", "Graceful degradation: deterministic workflow proceeds with the LLM disabled; AI text optional/never blocking; LLM feature flag; coordinators proceed manually", "Zensar Tech Lead"),
      ("LLM/Compliance", "PII / data-residency / audit-retention sign-off late (agent_runs holds vendor names + comments)", "Medium", "High", "Engage Legal/Privacy pre-Day-1; EU residency in DPA; retention policy + immutable audit trail agreed", "Shell Legal/Privacy / IT Security"),
      ("Scope/Schedule", "Scorecard taxonomy not locked", "Medium", "High", "Lock 4-category / 16-parameter taxonomy at the Day-2 design checkpoint", "Shell Product Owner / VMO lead"),
      ("Scope/Schedule", "Scorecard form-platform decision late", "Medium", "High", "Native in-app form recommended and pre-agreed; confirmed at Day-2 checkpoint", "Shell PO / Zensar SA"),
      ("Scope/Schedule", "Day-2 checkpoint produces too many in-scope changes (buffer exhausted)", "Medium", "High", "Strict change-control; route non-essentials to Phase 2 / warranty; formal re-baseline if breached", "Zensar Delivery Mgr / Shell PO"),
      ("Scope/Schedule", "Post-freeze (mid-Wk3, Day-13) scope changes", "Medium", "High", "Design frozen end of Wk1, code frozen mid-Wk3 (Day 13); later changes follow defect or Phase-2 path; change-control board", "Zensar Delivery Mgr / Shell PO"),
      ("Scope/Schedule", "One major UAT defect pushes cutover past its Wk4 slot", "Medium", "High", "P1 same-day triage; daily defect burn-down; CAB contingency slot held", "Zensar QA / Delivery Mgr"),
      ("Scope/Schedule", "Outlook email-rendering issues found late", "Medium", "Medium", "Table-based inline-style HTML + plain-text fallback; Litmus / Email-on-Acid QA in Wk2", "Zensar Frontend / QA"),
      ("Scope/Schedule", "Zensar team-member unplanned absence (no-PTO assumption)", "Medium", "High", "Cross-skilling; Tech Lead / Solution Architect cover; daily stand-up reprioritisation", "Zensar Delivery Mgr"),
      ("Scope/Schedule", "Late IT-Security objection blocking cutover", "Medium", "High", "Engage Security liaison from Day 1; design review early; Wk4 security sign-off planned", "Shell IT Security / Zensar SA"),
      ("Scope/Schedule", "No automated tests exist today (quality risk under compressed timeline)", "High", "Medium", "Risk-based manual UAT with 3 named coordinators; focus on deterministic logic; smoke tests; defects triaged in warranty", "Zensar QA"),
    ]
    for i, (cat, risk, like, imp, mit, owner) in enumerate(risks, 1):
        write_data_row(ws, r, c0, [i, cat, risk, like, imp, mit, owner])
        ws.cell(row=r, column=c0+3).alignment = CTR_TOP
        ws.cell(row=r, column=c0+4).alignment = CTR_TOP
        ws.row_dimensions[r].height = 34
        r += 1
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)

# --------------------------------------------------------------------------- acceptance criteria sheet
def build_acceptance_sheet(wb):
    ws = wb.create_sheet("Acceptance Criteria")
    ws.sheet_view.showGridLines = False
    cols = ["#", "Area", "Criterion", "Evidence of acceptance"]
    wid  = [4, 22, 50, 70]
    c0 = 2; set_widths(ws, c0, wid)
    title_block(ws, c0, len(cols), "Acceptance Criteria  -  Definition of Done for the pilot go-live (evidenced in Shell's tenant)")
    banner(ws, 4, c0, len(cols),
           "Verified jointly by the Zensar Tech Lead & Solution Architect with Shell's Product Owner (VMO lead) and IT Security liaison during Week-3 UAT and at the Week-4 production cutover. Evidence (demo / screenshot / log entry / sign-off) is captured per item and retained in the decision log committed to Shell's repository.",
           fill=SUB_FILL, font=Font(name=FONT, size=9, color="FF1F4E78"), height=28)
    hdr = 6; header(ws, hdr, c0, cols); r = hdr + 1
    groups = [
      ("Platform & Security", [
        ("Production environment live", "Running in Shell's Azure subscription on App Service (Linux containers) behind Front Door + WAF; nonprod & prod RGs in place."),
        ("Entra ID SSO", "A Shell user signs in via Entra ID (Azure AD) OIDC SSO; no local credentials exist."),
        ("RBAC enforced", "Four roles (VMO coordinator, VMO admin, executive sponsor, viewer) mapped to Entra security groups; correctly gate UI & API access."),
        ("Secrets management", "No secrets in code or .env; all secrets resolved from Key Vault via Managed Identity. POC secrets scrubbed from history and rotated."),
        ("Database", "Data persisted in Azure PostgreSQL Flexible Server (HA / zone-redundant in prod); JSON-file persistence retired."),
        ("Observability & audit", "App Insights + Log Analytics active; correlation IDs flow end-to-end; immutable audit trail records all actions & approvals; each outbound communication is attributable to the named approving coordinator and timestamped."),
        ("Single-tenant isolation & scope boundary", "Deployed solely within Shell's Azure subscription + M365 tenant; no Shell data stored outside Shell's boundary. Launch is English-only, web-only (desktop browser); multi-language UI, mobile/native clients and multi-tenancy are out of scope. Evidenced by deployment-topology + data-flow review at the Wk4 security sign-off."),
        ("Graceful degradation", "If the LLM provider is unavailable, the deterministic workflow continues; AI-assisted text is optional and never blocking; coordinators can proceed manually. LLM features are disable-by-configuration."),
        ("Security sign-off", "Shell IT Security review completed and signed off in Week 4."),
      ]),
      ("Functional (A-F)", [
        ("A - Scheduling & Coordination", "Roster refreshed; attendance outreach sent; candidate times via Graph findMeetingTimes against real mailboxes; slots ranked deterministically (organiser + sponsor free = hard constraint; soft scoring on attendance %, calendar conflicts, timezone fit); Teams meeting + invite via Graph after approval; RSVPs tracked."),
        ("B - Scorecard Input & Validation", "Personalised requests via Graph mail (approved before dispatch); responses via native in-app form (1-5 across 4 categories - Risk & Compliance, Performance, Commercial, Relationship - / 16 parameters); escalating reminders; deterministic validation (range, mandatory comment on 1/5, ~1.5-sigma outlier); internal-vs-vendor scorecard compiled."),
        ("C - Internal Alignment", "Score deltas vs previous cycle; divergence flagged where spread >= 1.5; LLM 'what changed' narrative + action-item extraction; internal-only Teams alignment meeting scheduled via Graph after approval."),
        ("D - Vendor Prep", "AI vendor brief (score/trend, per-category, concerns, positives, open actions); 3 response stances (Factual/Neutral/Escalation) per pushback; AI drafting suppressed for legal-review items; unresolved items tracked."),
        ("E - EGB/QBR Meeting", "Pre-meeting facilitator briefing; live notes across the 5 types; pasted transcript parsed into structured notes; formal minutes; action items extracted; approved minutes distributed via Graph mail. AI parses transcript + drafts briefing/minutes; 5-type classification, action-item register and state advance are deterministic."),
        ("F - Cross-Cycle Memory & Trends", "Multi-cycle trends; recurring-issue detection (category below threshold for >= 2 cycles); per-vendor trajectory; executive leadership brief; cross-vendor comparison. Read-only; any externally-shared brief follows the approval gate."),
        ("Workflow & gate invariants", "12-state forward-only machine behaves correctly (CYCLE_CREATED -> ATTENDEE_REFRESH_SENT -> AVAILABILITY_COLLECTED -> MEETING_SCHEDULED -> SCORECARD_REQUEST_SENT -> SCORECARD_COLLECTION -> SCORECARD_COMPILED -> INTERNAL_ALIGNMENT -> VENDOR_PREP -> MEETING_IN_PROGRESS -> POST_MEETING_COMPLETE -> ARCHIVED; invalid transitions rejected); human-approval gate enforced on every outbound action; deterministic-first / AI-second preserved (LLM only produces human-readable text)."),
      ]),
      ("Integration (Graph / Teams / Outlook)", [
        ("App-only certificate auth", "All email/calendar/Teams operations via Graph using app-only auth with a certificate (MSAL client-credentials); no user/delegated bearer token."),
        ("Service-mailbox sender", "Outbound mail sent from the dedicated service mailbox, constrained by the Exchange Application Access Policy."),
        ("Google stack removed", "Google / Gmail / Google Forms removed entirely from codebase and runtime."),
        ("Teams meetings", "Teams online meetings created via Graph for vendor-facing and internal-only meetings; vendor (external) attendees invitable per confirmed policy."),
        ("Outlook email rendering", "Table-based inline-style HTML renders acceptably on Outlook desktop, OWA & mobile, with plain-text fallback; verified via cross-client render QA."),
        ("Graph resilience", "Graph calls resilient under load (retry/backoff on 429 & 5xx)."),
      ]),
      ("Operations & Handover", [
        ("Admin module", "Vendor master, user-role view, LLM budget panel, audit-log viewer, system-health view all complete."),
        ("Branding", "Shell branding applied across the UI."),
        ("Observability & runbook", "Dashboards, alerts and runbook v1.0 in place; runbook handed to Shell IT Ops."),
        ("Training", "Coordinator training delivered to the 3 named VMO coordinators and recorded."),
        ("Documentation", "Architecture docs, decision log and supporting docs committed to Shell's repository."),
        ("LLM budget", "LLM operating within the agreed monthly token budget, with budget guards and rate limiting active."),
        ("Accessibility", "Automated accessibility checks (axe / Lighthouse) pass on the SPA at launch; a formal manual WCAG audit is deferred to warranty / Phase 2."),
      ]),
      ("Pilot", [
        ("Live pilot cycle", "One real, pre-agreed pilot vendor cycle progressing through the workflow in production with no open P1 issues."),
        ("UAT complete", "UAT completed with the 3 named VMO coordinators."),
        ("Defects closed", "All P1 & P2 UAT defects closed; P3 triaged and scheduled (warranty or Phase 2)."),
        ("Cutover", "Production cutover completed in Wk4 (DNS, slot-swap, 60-min observation) with CAB approval."),
      ]),
      ("Delivery Gates (all must pass in sequence)", [
        ("Design alignment / freeze", "Day-2 checkpoint locks ~12 decisions; design frozen end of Wk1; Module A demonstrated end-to-end on real Graph by Friday of Wk1."),
        ("Code freeze", "Mid Wk3 (Wed 8-Jul, Day 13): all six agents functional; scorecard collection in chosen mode; Shell-branded UI; Outlook templates; admin module; final hardening complete. From here: defects only."),
        ("UAT sign-off", "End of Wk3 (Day 15): UAT complete with all P1/P2 defects closed; release candidate approved. No production go-live this week."),
        ("Security sign-off", "Wk4: Shell IT Security review passed; observability, alerts and runbook in place."),
        ("CAB approval & cutover", "Wk4: CAB approval; production cutover (DNS, slot-swap, 60-min observation); pilot vendor cycle kicked off."),
        ("Handover", "Wk4: coordinator training delivered & recorded; runbook v1.0, documentation and decision log handed over."),
      ]),
      ("Warranty Terms", [
        ("4-week defect warranty", "After go-live, Zensar fixes P1/P2 defects (P1 same-day where feasible) while Shell IT Ops runs BAU; P3 scheduled into warranty backlog or Phase 2 by agreement."),
        ("Post-freeze change routing", "Any design/functional change requested after the Wk1 design freeze routes to the defect path or a Phase-2 change; not part of pilot acceptance."),
        ("Warranty dates", "Warranty start and end dates are agreed in writing as part of go-live sign-off."),
      ]),
    ]
    n = 1
    for area, items in groups:
        banner(ws, r, c0, len(cols), area, fill=SUB_FILL, font=Font(name=FONT, bold=True, size=10, color="FF1F4E78"), height=16)
        r += 1
        for crit, ev in items:
            write_data_row(ws, r, c0, [n, area, crit, ev])
            ws.cell(row=r, column=c0+2).font = BOLD_FONT
            ws.row_dimensions[r].height = 38
            r += 1; n += 1
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)


# =================================================================================== #
#  CLIENT PLAN  -  single clean sheet, manager-friendly, dependencies still shown      #
# =================================================================================== #
CCOLS = ["Sr", "Phase", "Activity", "Start", "Days", "End", "Depends On", "Needed By", "Outcome / Notes"]
CWID  = [4.5, 9, 50, 10.5, 6, 10.5, 26, 22, 56]

_PM  = d(2026,6,18)   # pre-mobilisation readiness window (Thu before T-0)
_PME = d(2026,6,19)
C0_LIST = [
 ("PHASE 0  -  Pre-Mobilisation  -  Shell readiness, confirmed in the 18-19 Jun window before the 22-Jun start (long-lead items initiated 4-6 weeks earlier)", [
   ("Identity & single sign-on", "Shell Entra admin", _PM, 2, _PME,
    "-", "HARD by T-0 (22-Jun)",
    "Entra app registrations with certificate credentials, OIDC single sign-on, and security groups for the 4 roles (VMO coordinator, admin, executive sponsor, viewer)."),
   ("Microsoft Graph admin consent  (LONG POLE)", "Shell IT Security + Entra", _PM, 2, _PME,
    "App registrations", "Initiated ~4-6 wks early; clears by end Wk1 (26-Jun)",
    "A 2-4 week Shell security review - the single biggest schedule risk - so it must be started weeks before mobilisation. Needed for Module A on real Graph by Friday of Week 1."),
   ("Service mailbox & least-privilege policy", "Shell Exchange admin", _PM, 2, _PME,
    "-", "HARD by T-0 (22-Jun)",
    "Dedicated vendorpulse-svc mailbox (E3/E5 licence) plus the Exchange Application Access Policy that confines the service to that one mailbox."),
   ("Teams & external-attendee policies", "Shell Teams / IT Security", _PM, 2, _PME,
    "Service mailbox", "By Wk1 (decision); pilot by Wk4",
    "Teams meeting-creation policy for the service mailbox, and confirmation it may invite external (vendor) attendees. Email-render QA account (Litmus) for Day-8 template checks."),
   ("Azure compute, data & secrets", "Shell Azure admin + DBA", _PM, 2, _PME,
    "Subscription", "HARD by T-0 (22-Jun)",
    "Subscription + nonprod/prod resource groups, App Service + Container Registry, PostgreSQL Flexible Server, and Key Vault - the platform the application is built on."),
   ("Networking, connectivity & egress", "Shell Networking", _PM, 2, _PME,
    "Subscription", "HARD by T-0 (DNS/TLS by Wk4)",
    "Front Door + WAF, private endpoints, the outbound egress allow-list to Graph + the AI endpoints, and App Insights. Public DNS + TLS are needed for the Week-4 cutover, not Day 1."),
   ("LLM provider contract + DPA  (LONG POLE)", "Shell Procurement + Legal", _PM, 2, _PME,
    "-", "Decided by T-0; initiated ~4-6 wks early; key for Wk2",
    "A 4-8 week contract/DPA - Anthropic Claude recommended (zero-retention, no-training, EU-residency); Azure OpenAI a config-only alternative. Powers the AI text in Modules C-F."),
   ("Source control, CI/CD & engineer access", "Shell IT / Zensar DevOps", _PM, 2, _PME,
    "Subscription", "HARD by T-0 (22-Jun)",
    "VPN / device-compliance access to the tenant, repo and non-prod Graph (a frequently-underestimated long pole), and the source-control + CI/CD platform with Zensar engineers onboarded as guests."),
   ("Zensar delivery tooling", "Zensar", _PM, 2, _PME,
    "-", "HARD by T-0 (22-Jun)",
    "Engineer laptops with a full local environment + admin rights, an approved IDE, and Claude Code / Claude for Work licences - the AI-assisted velocity the 4-week plan depends on."),
   ("Legal / Privacy sign-off path", "Shell Legal/Privacy", _PM, 2, _PME,
    "LLM DPA", "Go-live precondition (Wk4)",
    "Data-residency and audit-retention sign-off path agreed (the audit log holds vendor names + scorecard comments). Started early alongside the DPA."),
   ("People named & key decisions locked", "Shell + Zensar", _PM, 2, _PME,
    "-", "HARD by T-0 (22-Jun)",
    "Product Owner/VMO lead, liaisons, admins and 3 UAT coordinators named; pre-Day-1 decisions locked (provider, region, mailbox name + display name, database + CMK, CI/CD platform, vendor-master format, external-attendee policy)."),
 ]),
 ("PHASE 1  -  Week 1 (22-26 Jun)  -  Foundations, Migration & Design Alignment", [
   ("Secure the platform", "Zensar", W1[0], 2, d(2026,6,23),
    "Azure platform; CI/CD (P0)", "Days 1-2",
    "Branch the code and scrub POC secrets from history; wire Key Vault + Managed Identity so no secrets live in code or config. Day-1 connectivity smoke-tests."),
   ("Migrate data to Azure PostgreSQL", "Zensar", W1[0], 4, d(2026,6,25),
    "PostgreSQL provisioned (P0)", "Days 1-4",
    "Move the flat JSON / SQLite data into a managed PostgreSQL database (~15 tables) - the production system of record for workflow state, scores and audit."),
   ("Remove the Google / Gmail / Forms stack", "Zensar", W1[0], 3, d(2026,6,24),
    "Branch created", "Days 1-3",
    "The consumer Google stack used in the POC is removed entirely; everything moves to Microsoft 365 (Graph, Teams, Outlook)."),
   ("Connect Microsoft Graph (secure auth)", "Zensar", d(2026,6,23), 4, d(2026,6,26),
    "Admin consent; mailbox (P0)", "Days 2-5",
    "GraphService re-platformed onto app-only certificate authentication from the service mailbox - no pasted tokens."),
   ("Entra single sign-on & role-based access", "Zensar", d(2026,6,23), 4, d(2026,6,26),
    "Entra SSO + groups (P0)", "Days 2-5",
    "Users sign in with their Shell identity; the four roles are enforced on the server, never in the browser."),
   ("Wire the AI provider", "Zensar", d(2026,6,24), 2, d(2026,6,26),
    "LLM key (P0)", "Days 3-5",
    "Provider-abstracted LLM (Anthropic Claude recommended; Azure OpenAI a config-only switch). Agent layer is self-built tool-calling - no MAF SDK. AI is optional and never blocks the workflow."),
   ("Day-2 Design Alignment Checkpoint  [GATE]", "Shell + Zensar", d(2026,6,23), 1, d(2026,6,23),
    "Decisions locked (P0)", "GATE - Day 2 (23-Jun)",
    "Lock ~12 design decisions: scorecard mechanism & the 4-category/16-parameter taxonomy (Risk & Compliance, Performance, Commercial, Relationship), brief/minutes formats, vendor-master source, approval routing."),
   ("Module A live on real Graph", "Zensar", d(2026,6,24), 3, d(2026,6,26),
    "Graph admin consent; mailbox (P0)", "By Fri (Day 5)",
    "Scheduling & Coordination end-to-end on real Graph: roster refresh, find times, deterministic slot ranking, Teams meeting + invite, RSVP tracking."),
   ("Design freeze  [GATE]", "Shell + Zensar", W1[1], 1, W1[1],
    "Module A demo", "GATE - Day 5 (26-Jun)",
    "Scope frozen; later changes route to the defect or Phase-2 path."),
 ]),
 ("PHASE 2  -  Week 2 (29 Jun-3 Jul)  -  Functional Build", [
   ("Scorecard module (B)", "Zensar", W2[0], 4, d(2026,7,2),
    "Taxonomy locked (Day 2)", "Days 6-10",
    "Native in-app scorecard form (1-5 across 4 categories / 16 parameters); deterministic validation and outlier flagging; requests + escalating reminders via Outlook. Approval gate on the emails."),
   ("Alignment (C) & Vendor Prep (D)", "Zensar", W2[0], 4, d(2026,7,2),
    "LLM key; Graph (P0)", "Days 6-10",
    "Score-change narrative and action items; internal Teams alignment meeting; AI vendor brief and three response stances per pushback (AI suppressed on legal-review items)."),
   ("Meeting (E) & Cross-Cycle Trends (F)", "Zensar", W2[0], 4, d(2026,7,2),
    "LLM key; Postgres", "Days 6-10",
    "Facilitator briefing, transcript-to-minutes, action items, and minutes distribution via Outlook; multi-cycle trends, recurring-issue detection and the executive leadership brief."),
   ("Shell-branded user interface", "Zensar", W2[0], 3, d(2026,7,1),
    "Brand assets (P0)", "Days 6-8",
    "Shell branding applied across the web app; the approval panel is the primary review surface for every AI draft."),
   ("Outlook email templates + render QA", "Zensar", W2[0], 4, d(2026,7,2),
    "Brand sign-off by Day 8 (P0)", "Day 8 (brand) / Day 10",
    "Outlook-friendly templates verified on Outlook desktop, OWA and mobile via cross-client QA."),
   ("Admin module, audit trail & accessibility", "Zensar", W2[0], 4, d(2026,7,2),
    "App Insights (P0)", "Days 6-10",
    "Admin views (vendor master, roles, AI budget, audit-log viewer, system health); append-only audit trail with approver attribution; automated accessibility checks."),
   ("Initial hardening", "Zensar", d(2026,7,1), 3, d(2026,7,3),
    "All modules", "Days 8-10",
    "Rate limiting, AI token-budget guards, retry/backoff and end-to-end correlation IDs (completed before the mid-Wk3 code freeze)."),
   ("All six agents functional  (milestone)", "Zensar", W2[1], 1, W2[1],
    "Modules A-F", "End of Wk2 (3-Jul)",
    "Functional build substantially complete; remaining completion + final hardening carry into early Wk3, up to the Day-13 code freeze."),
 ]),
 ("PHASE 3  -  Week 3 (6-10 Jul)  -  Completion, mid-week Code Freeze (Wed 8-Jul), Stabilisation & UAT (no go-live)", [
   ("Final completion & hardening", "Zensar", W3[0], 3, d(2026,7,8),
    "Functional build (Wk2)", "Days 11-13 (by code freeze)",
    "Close out any remaining build and finish hardening ahead of the mid-week code freeze."),
   ("UAT with 3 named VMO coordinators", "Zensar + Shell", W3[0], 5, W3[1],
    "Modules functional; coordinators (P0)", "Days 11-15",
    "Risk-based testing across all six modules; begins on completed modules and intensifies after the code freeze (defect fixes only thereafter)."),
   ("Defect triage & daily burn-down", "Zensar", W3[0], 5, W3[1],
    "UAT", "Days 11-15",
    "P1 resolved same-day; P1/P2 closed before sign-off; P3 triaged to warranty/Phase 2."),
   ("Observability, alerting & runbook v1.0", "Zensar", W3[0], 5, W3[1],
    "App Insights (P0)", "Days 11-15",
    "Operational readiness for handover to Shell IT Ops."),
   ("Code freeze  [GATE]", "Zensar", d(2026,7,8), 1, d(2026,7,8),
    "All modules functional; hardening done", "GATE - Day 13 (Wed 8-Jul)",
    "Middle of Week 3. No further feature development; stabilisation & defect fixes only from here."),
   ("Regression & smoke testing", "Zensar", d(2026,7,8), 3, W3[1],
    "Code freeze", "Days 13-15 (post-freeze)",
    "Confirms the deterministic core (slot ranking, score validation, workflow transitions) behaves identically after re-platforming."),
   ("Validate platform capabilities", "Zensar + Shell", d(2026,7,8), 3, W3[1],
    "Audit; RBAC; Graph auth", "Days 13-15 (post-freeze)",
    "Audit append-only behaviour, role-based access, secure Graph auth and email rendering all verified before cutover."),
   ("UAT sign-off  [GATE]", "Shell + Zensar", W3[1], 1, W3[1],
    "Defects closed", "GATE - Day 15 (10-Jul)",
    "All P1/P2 UAT defects closed; release candidate approved for deployment."),
 ]),
 ("PHASE 4  -  Week 4 (13-17 Jul)  -  Security Sign-off, Deployment & Defect-Warranty", [
   ("Shell IT Security review & sign-off  [GATE]", "Shell IT Security", W4[0], 2, d(2026,7,14),
    "UAT sign-off", "GATE - early Wk4",
    "Sign-off on Graph scopes, Application Access Policy, RBAC and audit posture before cutover."),
   ("Confirm production prerequisites", "Shell + Zensar", W4[0], 2, d(2026,7,14),
    "DNS/TLS, CA path, Front Door (P0)", "Before cutover",
    "Production DNS + TLS, Conditional Access path for the service principal, Front Door/WAF prod config; pilot vendor-master (CSV/Excel) imported."),
   ("CAB approval & production cutover  [GATE]", "Shell CAB + Zensar", d(2026,7,14), 2, d(2026,7,15),
    "Security sign-off", "GATE - Wk4 (Wed-Thu)",
    "Change Advisory Board approval; go-live via DNS change + slot-swap with a 60-minute observation window and tested rollback."),
   ("Pilot cycle kick-off & handover  [GATE]", "Zensar + Shell", d(2026,7,16), 2, W4[1],
    "Cutover; vendor-master", "Wk4 (Thu-Fri)",
    "One real pilot vendor cycle kicked off in production (Tech Lead pair-shadowing); recorded coordinator training; runbook v1.0 + decision log handed over."),
   ("Defect-warranty support begins", "Zensar + Shell IT Ops", W4[1], 1, W4[1],
    "Go-live", "From go-live",
    "4-week warranty: Zensar fixes P1/P2 (P1 same-day where feasible) while Shell runs BAU; warranty dates agreed in writing at handover."),
 ]),
]

def build_client(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Implementation Plan"
    ws.sheet_view.showGridLines = False
    c0 = 2; ncol = len(CCOLS)
    set_widths(ws, c0, CWID)
    title_block(ws, c0, ncol, "CLIENT Implementation Plan  -  4-week delivery overview, dependencies & key dates")
    cinfo = Font(name=FONT, size=9, color="FF1F4E78")
    banner(ws, 4, c0, ncol,
           "Reading guide:  Phase 0 is Shell-side readiness, confirmed in the 18-19 Jun window before the Mon 22-Jun (T-0) start - 'HARD by T-0' items must be ready on the start date.  The two long-lead items (Graph admin consent, LLM contract/DPA) must be INITIATED ~4-6 weeks earlier.  Gates are shown in gold; orange = hard Day-1 prerequisite.",
           fill=SUB_FILL, font=cinfo, height=28)
    banner(ws, 5, c0, ncol,
           "Team (Zensar): 1 Delivery Mgr, 1 Solution Architect, 1 Tech Lead, 2 Backend, 2 Frontend, 1 QA, 1 DevOps - 100% allocated, no planned leave Wk1-3.   Scope basis: the proven POC behaviour (Modules A-F) is the baseline; this engagement re-platforms it onto Shell-grade Azure - it does not rebuild the features. New functional requirements are out of scope and re-estimated.",
           fill=SUB_FILL, font=cinfo, height=28)
    hdr = 7; header(ws, hdr, c0, CCOLS); r = hdr + 1; sr = 1
    for phase_title, rows in C0_LIST:
        banner(ws, r, c0, ncol, phase_title); r += 1
        phase_short = phase_title.split("  -  ")[0].replace("PHASE ", "P")
        for (act, owner, start, days, end, dep, needed, notes) in rows:
            is_gate = "[GATE]" in act
            is_block = isinstance(needed, str) and "HARD" in needed
            fill = GATE_FILL if is_gate else (BLOCK_FILL if is_block else None)
            vals = [sr, phase_short, act, start, days, end, dep, needed, notes]
            write_data_row(ws, r, c0, vals, fill=fill)
            if is_gate:
                ws.cell(row=r, column=c0+2).font = BOLD_FONT
            ws.row_dimensions[r].height = 44
            r += 1; sr += 1
    # milestones strip
    r += 1
    banner(ws, r, c0, ncol, "Key Dates & Gates", fill=SUB_FILL, font=Font(name=FONT, bold=True, size=10, color="FF1F4E78"), height=16); r += 1
    ms = [
      ("T-0 / Day 1", "Mon 22-Jun-2026", "Engagement start - all HARD Day-1 prerequisites in place"),
      ("Day-2 checkpoint [GATE]", "Tue 23-Jun", "~12 design decisions locked"),
      ("Design freeze [GATE]", "Fri 26-Jun", "Module A live on real Graph; scope frozen"),
      ("All six agents functional", "Fri 3-Jul", "End of Week 2 build milestone (not a gate)"),
      ("Code freeze [GATE]", "Wed 8-Jul", "Middle of Week 3; no further feature development - stabilisation only"),
      ("UAT sign-off [GATE]", "Fri 10-Jul", "P1/P2 defects closed; release candidate approved"),
      ("Security sign-off [GATE]", "Mon-Tue 13-14 Jul", "Shell IT Security approval"),
      ("CAB approval & cutover [GATE]", "Wed-Thu 15-16 Jul", "Go-live via DNS + slot-swap; pilot cycle kick-off"),
      ("Handover [GATE]", "Fri 17-Jul", "Training, runbook v1.0, decision log handed over; warranty begins"),
    ]
    msh = ["Milestone", "Date", "Meaning"]
    for i,h in enumerate(["", "", ""]):
        pass
    # mini header
    for i,h in enumerate(["Milestone","Date","Meaning"]):
        cell = ws.cell(row=r, column=c0+i, value=h); cell.fill=HEADER_FILL; cell.font=HEAD_FONT; cell.alignment=CTR_MID; cell.border=BORDER
    ws.merge_cells(start_row=r, start_column=c0+2, end_row=r, end_column=c0+ncol-1)
    r += 1
    for milestone, date, meaning in ms:
        ws.cell(row=r, column=c0, value=milestone).font=BOLD_FONT
        ws.cell(row=r, column=c0+1, value=date).font=CELL_FONT
        ws.merge_cells(start_row=r, start_column=c0+2, end_row=r, end_column=c0+ncol-1)
        ws.cell(row=r, column=c0+2, value=meaning).font=CELL_FONT
        for i in range(ncol):
            cc = ws.cell(row=r, column=c0+i); cc.border=BORDER; cc.alignment=WRAP_TOP
            if "GATE" in milestone: cc.fill = GATE_FILL
        r += 1

    # ---- Out of Scope + Assumptions (manager context, full-width bullet rows)
    def block(title, items):
        nonlocal r
        r += 1
        banner(ws, r, c0, ncol, title, fill=SUB_FILL, font=Font(name=FONT, bold=True, size=10, color="FF1F4E78"), height=16); r += 1
        for it in items:
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0+ncol-1)
            cell = ws.cell(row=r, column=c0, value="-  " + it)
            cell.font = CELL_FONT; cell.alignment = WRAP_TOP
            for i in range(ncol):
                ws.cell(row=r, column=c0+i).border = BORDER
            ws.row_dimensions[r].height = 26
            r += 1

    block("Out of Scope (this release)", [
        "Shell-wide vendor master / SAP Ariba / Coupa integration - the pilot vendor list is seeded from CSV/Excel; system integration is a follow-on phase.",
        "Contract-management feature - contract terms are referenced for context only, not stored or analysed.",
        "Finance / spend / SAP integration - spend figures are referenced manually within scorecards.",
        "Mobile native app (web-only, desktop browser) and multi-language UI (English only at launch).",
        "Microsoft Teams transcript ingestion - transcripts are pasted in for parsing at launch.",
        "Deferred to warranty / Phase 2: performance & load testing, third-party penetration test, formal manual WCAG audit, comprehensive automated regression suite, multi-vendor onboarding, cross-region DR, customer-managed keys (CMK).",
    ])
    block("Assumptions (the plan holds only if...)", [
        "Shell completes all pre-mobilisation access before Day 1; the two long-lead items (Graph admin consent, LLM contract/DPA) are initiated 4-6 weeks ahead, during contract negotiation.",
        "The Zensar team is 100% allocated with no planned leave across Weeks 1-3 (deployment & warranty in Week 4).",
        "AI-assisted delivery tooling (laptops + admin rights, approved IDE, Claude Code licences) is available to every engineer from Day 1 - the velocity the 4-week plan depends on.",
        "The proven POC behaviour (Modules A-F) is accepted as the baseline; this engagement re-platforms it, it does not rebuild the features. New functional requirements are out of scope and re-estimated.",
        "One pilot vendor cycle, with its vendor-master data provided in time for the Week-4 kick-off; the service mailbox may invite external (vendor) attendees.",
        "Design is frozen at the end of Week 1 and code in the middle of Week 3 (Wed 8-Jul); later changes follow the defect or Phase-2 path.",
        "Launch is English-only, web-only (desktop browser) and single-tenant (Shell only); platform-managed encryption keys (CMK deferred).",
    ])
    ws.freeze_panes = ws.cell(row=hdr+1, column=c0)
    wb.save(path)


if __name__ == "__main__":
    import os
    base = os.path.dirname(os.path.abspath(__file__))
    detailed = os.path.join(base, "VendorPulse_Implementation_Plan_Detailed.xlsx")
    client   = os.path.join(base, "VendorPulse_Implementation_Plan_Client.xlsx")
    n = build_detailed(detailed)
    build_client(client)
    print("Detailed task rows:", n)
    print("Wrote:", detailed)
    print("Wrote:", client)
