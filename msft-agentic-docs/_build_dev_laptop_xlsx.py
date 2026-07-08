"""Build VendorPulse_Developer_Laptop_Software.xlsx

Single sheet, two sections for the client to ready developer laptops:
  A. Requires Administrator Rights (IT to install / elevate)
  B. No Administrator Rights Required (developer self-install)
Plus a Notes & Assumptions block (prerequisites and likely changes).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "VendorPulse_Developer_Laptop_Software.xlsx"

NAVY = "1F3864"      # Section A banner
TEAL = "0B6E63"      # Section B banner
GREY = "33475B"      # column header
ZEBRA = "F2F5F8"
AMBER = "FFF4E5"
INK = "1A1A1A"
WHITE = "FFFFFF"

thin = Side(style="thin", color="C9D2DC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CT = Alignment(horizontal="center", vertical="top")

HEADERS = ["#", "Software / Component", "Recommended Version", "Purpose",
           "Install Method / Source", "Notes (admin alternative · licensing · scope)"]
WIDTHS = [4, 30, 24, 34, 34, 58]

# ── Section A: requires admin ────────────────────────────────────────────────
SECTION_A = [
    ["Python", "3.11.x (64-bit)", "Backend runtime — FastAPI, agents, AI Service",
     "python.org installer (all-users) or winget",
     "Standard all-users install needs admin. No-admin option: per-user install (untick “Install for all users”) or Microsoft Store."],
    ["Node.js", "20 LTS (≥ 20.19) or 22 LTS (≥ 22.12)", "Frontend build/runtime — Vite 8, React 19",
     "nodejs.org MSI (all-users)",
     "Vite 8 requires Node 20.19+/22.12+. No-admin option: nvm-windows or per-user install."],
    ["Git", "2.40+", "Version control",
     "Git for Windows installer (all-users)",
     "No-admin option: portable Git or per-user install."],
    ["Docker Desktop", "24+ (latest)", "Container image builds & local parity",
     "Docker Desktop installer (enables WSL2/Hyper-V)",
     "Needs admin + WSL2. Enterprise licence applies at Shell scale — confirm licence or approved alternative (Rancher Desktop / Podman). Optional day-to-day: CI builds the images."],
    ["WSL2 (Ubuntu)", "latest", "Linux subsystem (Docker backend; optional dev shell)",
     "wsl --install (Windows feature)",
     "Admin required to enable the Windows feature. Only needed if Docker Desktop is used."],
    ["Azure CLI (az)", "2.60+", "Azure / Foundry / Microsoft Graph auth & deployment",
     "MSI installer",
     "No-admin option: pip install azure-cli inside the project venv."],
    ["[Optional] JetBrains PyCharm / WebStorm", "latest", "Alternative IDE to VS Code",
     "JetBrains installer (or Toolbox)",
     "VS Code (Section B) is the default and needs no admin. JetBrains Toolbox installs per-user (no admin)."],
    ["[Optional] DBeaver / pgAdmin", "latest", "PostgreSQL database client",
     "Desktop installer",
     "No-admin option: DBeaver portable build (listed in Section B)."],
    ["[Optional] PowerShell 7", "latest", "Modern shell (CI/script parity)",
     "MSI installer",
     "Windows PowerShell 5.1 is preinstalled; PS7 is optional."],
]

# ── Section B: no admin ──────────────────────────────────────────────────────
SECTION_B = [
    ["Visual Studio Code (User Installer)", "latest", "Primary IDE",
     "code.visualstudio.com → “User Installer”",
     "Installs to %LOCALAPPDATA% — no admin. (The “System Installer” would need admin.)"],
    ["VS Code extensions", "latest", "IDE tooling (Python, web, Azure, AI)",
     "Installed inside VS Code (Marketplace)",
     "Python, Pylance, Ruff, ESLint, Prettier, Tailwind CSS IntelliSense, GitHub Copilot + Copilot Chat, Docker, Bicep, Azure Resources, GitLens."],
    ["Python packages (pip, in venv)", "per requirements.txt", "Backend libraries",
     "python -m venv .venv  +  pip install -r requirements.txt",
     "fastapi, uvicorn, pydantic, httpx, openai, azure-ai-projects, azure-identity, msal, SQLAlchemy, Alembic, psycopg/asyncpg, ruff, python-dotenv, truststore. Needs Python (Section A)."],
    ["Node packages (npm, project-local)", "per package.json", "Frontend libraries & tooling",
     "npm install  (in the frontend folder)",
     "react, react-dom, vite, typescript, zustand, recharts, tailwindcss, eslint, prettier. Needs Node (Section A)."],
    ["Terraform", "1.x (latest)", "Infrastructure as code",
     "Download binary → add to user PATH",
     "Single portable executable — no installer, no admin. (Choose Terraform or Bicep per Shell standard.)"],
    ["Bicep CLI", "latest", "Infrastructure as code (Azure-native)",
     "az bicep install (to user profile)",
     "No admin (needs Azure CLI). Standalone exe also available."],
    ["[Optional] GitHub CLI (gh)", "latest", "Git/GitHub operations from terminal",
     "Scoop / portable zip",
     "Portable/scoop needs no admin; the MSI would need admin."],
    ["[Optional] Trivy", "latest", "Image / dependency security scan (local)",
     "Portable binary → user PATH",
     "Runs in CI; local use optional. No admin."],
    ["[Optional] GitLeaks", "latest", "Secret scanning (local)",
     "Portable binary → user PATH",
     "Runs in CI; local use optional. No admin."],
    ["[Optional] DBeaver (portable)", "latest", "PostgreSQL database client",
     "Portable zip → run from user folder",
     "No-admin alternative to the desktop installer."],
]

NOTES = [
    "Versions are indicative and are pinned in requirements.txt / package.json / package-lock.json (scanned in CI). Treat ‘+’ / ≥ as a floor.",
    "Section B ‘no-admin’ packages (pip / npm) require the runtimes in Section A (Python, Node) to be installed first.",
    "Corporate proxy / firewall must allow: PyPI (pypi.org), npm registry (registry.npmjs.org), VS Code Marketplace, GitHub, and Azure/Microsoft Graph endpoints — otherwise installs fail behind the proxy.",
    "AI pair-programming tool: GitHub Copilot (licence required). Anthropic / Claude tooling is out of scope.",
    "No local GPU is required — all model inference runs remotely in Microsoft Foundry.",
    "Likely-to-change items: Docker (Desktop vs Rancher/Podman), Node 20 vs 22, IaC (Terraform vs Bicep), and DB client — align to Shell standards during setup.",
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Developer Laptop Setup"

    # title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = "VendorPulse — Developer Laptop Software Checklist"
    c.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws.merge_cells("A2:F2")
    ws["A2"].value = ("Prepared by Zensar for Shell  ·  for provisioning developer laptops  ·  two sections: "
                      "(A) needs admin rights, (B) no admin rights")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws.merge_cells("A3:F3")
    ws["A3"].value = ("Items marked [Optional] are not strictly required for the MVP build. ‘No-admin’ alternatives "
                      "are noted where they exist.")
    ws["A3"].font = Font(name="Calibri", size=9.5, color="555555")

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5

    def banner(text, fill):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

    def header():
        nonlocal row
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = PatternFill("solid", fgColor=GREY)
            cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="center" if ci == 1 else "left")
            cell.border = BORDER
        row += 1

    def rows(data):
        nonlocal row
        for idx, item in enumerate(data, 1):
            vals = [idx] + item
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = Font(name="Calibri", size=9.5, color=INK,
                                 bold=(ci == 2))
                cell.alignment = CT if ci == 1 else WRAP
                cell.border = BORDER
                if idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=ZEBRA)
            row += 1

    banner("A.  Requires Administrator Rights  (IT to install / elevate)", NAVY)
    header()
    rows(SECTION_A)
    row += 1
    banner("B.  No Administrator Rights Required  (developer can self-install)", TEAL)
    header()
    rows(SECTION_B)

    # notes block
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    nb = ws.cell(row=row, column=1, value="Notes & Assumptions")
    nb.fill = PatternFill("solid", fgColor="B45309")
    nb.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    nb.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1
    for n in NOTES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="•  " + n)
        cell.font = Font(name="Calibri", size=9.5, color=INK)
        cell.alignment = WRAP
        cell.fill = PatternFill("solid", fgColor=AMBER)
        cell.border = BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    # print setup: landscape, fit to one page wide
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"
    ws.freeze_panes = "A5"

    wb.save(OUT)
    print("Wrote", OUT, "| Section A:", len(SECTION_A), "items | Section B:", len(SECTION_B), "items")


if __name__ == "__main__":
    build()
