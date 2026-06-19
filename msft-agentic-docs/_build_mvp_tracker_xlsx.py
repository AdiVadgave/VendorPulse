"""Build VendorPulse_MVP_Delivery_Tracker.xlsx

A living progress tracker for the VendorPulse MVP build. Sheets:
  1. Overview        — how to use, legend, sheet index
  2. MVP Scope       — what's in / out of the MVP baseline
  3. Milestones      — phase/week plan with status
  4. Task Tracker    — granular tasks with Definition of Done (the main sheet)
  5. Module Exit Criteria — the gate checklist that must pass to call a module done
  6. Weekly Progress Log  — running log to track velocity / avoid drift
  7. Scope Change Log     — discovery-phase changes vs the MVP baseline
  8. Dashboard       — auto summaries (formulas)

Usage: python _build_mvp_tracker_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

OUT = "VendorPulse_MVP_Delivery_Tracker.xlsx"

# palette
NAVY = "1F3A5F"
RED = "C8102E"
GREY = "E9ECEF"
WHITE = "FFFFFF"
INK = "1A1A1A"
# status colours
C_DONE = "C6EFCE"
C_PROG = "FFEB9C"
C_BLOCK = "FFC7CE"
C_REVIEW = "BDD7EE"
C_NOT = "EDEDED"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
TFONT = Font(name="Calibri", size=10, color=INK)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
HFILL = PatternFill("solid", fgColor=NAVY)
AREAFILL = PatternFill("solid", fgColor=GREY)

STATUSES = ["Not Started", "In Progress", "In Review", "Blocked", "Done"]


def header_row(ws, row, headers, fill=HFILL):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HFONT
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, wrap_cols=(), bold=False, fill=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name="Calibri", size=10, bold=bold, color=INK)
        cell.alignment = WRAP if (c in wrap_cols) else TOP
        cell.border = BORDER
        if fill:
            cell.fill = fill


def title_block(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=RED)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")


# ----------------------------------------------------------------------------
# Task data: (Area, ID, Task/Deliverable, Definition of Done, Est days)
# ----------------------------------------------------------------------------
TASKS = [
    # --- Foundations (cross-cutting, mostly Week 1 + ongoing) ---
    ("Foundations", "FND-1", "Provision 2 Azure VMs in Private VNet (IaC: Bicep/Terraform)",
     "Both VMs deployed via IaC; VNet-private; reproducible; peer-reviewed", 2),
    ("Foundations", "FND-2", "Application Gateway + WAF (OWASP), origin-lock, TLS",
     "Only App Gateway reaches the VMs; OWASP rules on; TLS terminated; backend has no public IP", 1),
    ("Foundations", "FND-3", "PostgreSQL Flexible Server (VNet-private, SSL)",
     "DB reachable only inside VNet; SSL enforced; credentials stored in Key Vault", 1),
    ("Foundations", "FND-4", "JSON/SQLite -> PostgreSQL migration (BaseRepository seam)",
     "All entities persist to Postgres; repository seam swapped; no file persistence; data integrity verified", 2),
    ("Foundations", "FND-5", "Key Vault + Managed Identity",
     "All secrets (LLM key, Graph cert, JWT key, DB creds) in KV; fetched via MI at runtime; no secrets in code/.env in deployed envs", 1),
    ("Foundations", "FND-6", "Entra SSO (OIDC) + JWT validation",
     "Login via Shell Entra; every protected route validates JWT against well-known config; unauthenticated requests rejected", 1.5),
    ("Foundations", "FND-7", "RBAC (Coordinator / Sponsor / Viewer) from token claims",
     "Roles enforced server-side; UI reflects role; Viewer cannot mutate state", 1),
    ("Foundations", "FND-8", "Graph app-only certificate auth (MSAL, mailbox-scoped)",
     "App authenticates with cert from KV; Application Access Policy limits to service mailbox; token auto-refresh", 1.5),
    ("Foundations", "FND-9", "Remove Google / Gmail / Forms components",
     "No Google packages or config remain; functions moved to Graph + native form", 0.5),
    ("Foundations", "FND-10", "App Insights + OpenTelemetry instrumentation",
     "Traces/metrics/logs emitted via OTel; correlation IDs link HTTP <-> service <-> model; immutable audit trail", 1),
    ("Foundations", "FND-11", "Egress proxy configuration on VM2 (HTTP(S)_PROXY)",
     "All outbound (Graph, Foundry) routed via Shell egress; no direct internet egress", 0.5),
    ("Foundations", "FND-12", "WorkflowEngine (12-state, forward-only) verified/ported",
     "Invalid transitions raise error -> 409; single source of truth; unit-tested across all 12 states", 1),
    ("Foundations", "FND-13", "Approval gate (HITL) framework",
     "AI outputs flagged requires_approval; external actions fire only from deterministic route post-approval; agent has no side-effect tool", 1.5),
    ("Foundations", "FND-14", "AI Service (Azure AI Foundry GPT-4o; MAF removed)",
     "AI Service wraps Foundry via Azure OpenAI SDK; in-tenant; returns drafted text only; structured output validated", 1.5),
    ("Foundations", "FND-15", "AgentResponse contract + adapter",
     "Single response shape for AI and deterministic paths; frontend depends on it; both paths emit identical envelope", 1),
    ("Foundations", "FND-16", "CI/CD pipeline (build, lint, SAST/dep/secret scan, deploy) + regression tests",
     "Pipeline gates deploys; scans pass; regression covers approval gate, deterministic path, AgentResponse; prod needs manual approval", 2),

    # --- Module A: Scheduling ---
    ("Module A — Scheduling", "A1", "Cycle creation (route + repo, state CYCLE_CREATED)",
     "Cycle created, persisted; state set via WorkflowEngine", 1),
    ("Module A — Scheduling", "A2", "Attendee refresh (Graph User.Read.All)",
     "Attendee list pulled/refreshed; state ATTENDEE_REFRESH_SENT", 1),
    ("Module A — Scheduling", "A3", "Availability collection (Graph calendar free/busy)",
     "Availability gathered for attendees; state AVAILABILITY_COLLECTED", 1),
    ("Module A — Scheduling", "A4", "Deterministic slot ranking",
     "Slots ranked by deterministic algorithm; reproducible; no LLM", 1),
    ("Module A — Scheduling", "A5", "AI Service: polish invite text (optional)",
     "Draft invite generated; requires_approval; deterministic fallback text when LLM off", 0.5),
    ("Module A — Scheduling", "A6", "Approval + send invites (Graph Calendars.ReadWrite, OnlineMeetings)",
     "Invite/Teams meeting created only after human approval, from deterministic route; state MEETING_SCHEDULED", 1.5),
    ("Module A — Scheduling", "A7", "RSVP status tracking",
     "RSVP statuses read from Graph and surfaced in UI", 0.5),
    ("Module A — Scheduling", "A8", "Frontend Scheduling module + ApprovalPanel + progress bar",
     "UI drives Module A end to end; role-gated; reflects workflow state", 1.5),

    # --- Module B: Scorecard ---
    ("Module B — Scorecard", "B1", "Scorecard request distribution (native form magic-link via Graph mail)",
     "Requests sent to stakeholders; state SCORECARD_REQUEST_SENT", 1),
    ("Module B — Scorecard", "B2", "Native in-app scorecard form (React) + SSO / one-time token",
     "Internal via SSO, external via one-time token; submissions validated", 1.5),
    ("Module B — Scorecard", "B3", "Collection + polling / status",
     "Submissions tracked; state SCORECARD_COLLECTION", 0.5),
    ("Module B — Scorecard", "B4", "Deterministic validation",
     "All scores validated by rules; invalid rejected; no LLM", 1),
    ("Module B — Scorecard", "B5", "Compile + outlier detection (deterministic)",
     "Compiled scorecard; outliers flagged deterministically; state SCORECARD_COMPILED", 1),
    ("Module B — Scorecard", "B6", "Frontend Scorecard module + status dashboard",
     "Coordinator sees collection status; compiled view renders", 1),

    # --- Module C: Alignment ---
    ("Module C — Alignment", "C1", "AI Service: extract action items from notes",
     "Draft action items; requires_approval; grounded in notes; audited", 1),
    ("Module C — Alignment", "C2", "AI Service: summarise score changes",
     "Draft summary grounded in compiled scorecard; requires_approval", 1),
    ("Module C — Alignment", "C3", "Approval gate integration",
     "AI outputs approved before use; deterministic fallback available", 0.5),
    ("Module C — Alignment", "C4", "Frontend Alignment module",
     "UI shows/edits/approves; state INTERNAL_ALIGNMENT", 1),

    # --- Module D: Vendor Prep ---
    ("Module D — Vendor Prep", "D1", "AI Service: generate vendor brief",
     "Draft brief grounded in cycle data; requires_approval; audited", 1),
    ("Module D — Vendor Prep", "D2", "AI Service: draft 3 response options per pushback",
     "3 distinct options per pushback; requires_approval", 1),
    ("Module D — Vendor Prep", "D3", "Approval gate integration",
     "Brief and responses approved before they can be shared", 0.5),
    ("Module D — Vendor Prep", "D4", "Frontend Vendor Prep module",
     "UI renders brief + options; state VENDOR_PREP", 1),

    # --- Module E: Meeting ---
    ("Module E — Meeting", "E1", "AI Service: parse transcript",
     "Transcript parsed into structured notes; requires_approval", 1),
    ("Module E — Meeting", "E2", "AI Service: generate meeting minutes",
     "Draft minutes; requires_approval; grounded in transcript", 1),
    ("Module E — Meeting", "E3", "Blob storage for transcripts / minutes (immutable)",
     "Stored append-only; retrievable; never updated/deleted", 0.5),
    ("Module E — Meeting", "E4", "Approval + send minutes (Graph Mail.Send)",
     "Minutes sent only after approval; state MEETING_IN_PROGRESS -> POST_MEETING_COMPLETE", 1),
    ("Module E — Meeting", "E5", "Frontend Meeting module",
     "UI captures meeting, shows minutes, drives state", 1),

    # --- Module F: Analytics ---
    ("Module F — Analytics", "F1", "Deterministic analytics (vendor trajectories / trends)",
     "Computed deterministically from cycle history; reproducible", 1),
    ("Module F — Analytics", "F2", "AI Service: leadership brief card",
     "Draft brief card; requires_approval; audited", 1),
    ("Module F — Analytics", "F3", "Frontend Analytics page + charts (Recharts)",
     "Dashboards render per completed cycle; charts correct", 1),

    # --- Admin & branding ---
    ("Admin & Branding", "ADM-1", "Admin module (user/role management, cycle config)",
     "Admin can manage roles/config; RBAC-gated to admin role", 1.5),
    ("Admin & Branding", "ADM-2", "Shell branding applied",
     "Branding consistent across all UI via Design System", 0.5),
    ("Admin & Branding", "ADM-3", "Outlook email templates configured",
     "Templates used for invites and minutes; render correctly in Outlook", 0.5),

    # --- Deployment & go-live ---
    ("Deployment & Go-Live", "DEP-1", "Staging environment + smoke tests",
     "Staging deployed; automated smoke tests pass", 1),
    ("Deployment & Go-Live", "DEP-2", "Hardening & stabilisation",
     "UAT defects fixed; performance acceptable; error handling verified", 2),
    ("Deployment & Go-Live", "DEP-3", "UAT with Shell VMO team",
     "All UAT scenarios executed and signed off by Shell", 2),
    ("Deployment & Go-Live", "DEP-4", "Observability dashboards + runbook",
     "Dashboards live in App Insights; operational runbook documented", 1),
    ("Deployment & Go-Live", "DEP-5", "IT Security sign-off",
     "Shell IT Security review passed", 1),
    ("Deployment & Go-Live", "DEP-6", "CAB approval",
     "Change Advisory Board approval obtained", 0.5),
    ("Deployment & Go-Live", "DEP-7", "Production cutover & go-live",
     "Prod live; pilot vendor cycle kicked off", 1),
    ("Deployment & Go-Live", "DEP-8", "Coordinator training & handover",
     "Training delivered; handover docs complete", 1),
    ("Deployment & Go-Live", "DEP-9", "Defect-warranty support",
     "Warranty period active; defect process agreed", 0),

    # --- Compliance (Shell, parallel — has external lead time) ---
    ("Compliance (Shell)", "CMP-1", "AI Registry + ServiceNow registration",
     "System registered in AI Registry / ServiceNow", 0),
    ("Compliance (Shell)", "CMP-2", "IRM risk assessment / IAQ",
     "IRM IAQ completed and accepted", 0),
    ("Compliance (Shell)", "CMP-3", "EU AI Act classification",
     "Risk classification completed and recorded", 0),
    ("Compliance (Shell)", "CMP-4", "Shell.AI + TRB model/design approval",
     "Model and design approved by Shell.AI and TRB", 0),
    ("Compliance (Shell)", "CMP-5", "Entra app registrations (NonProd+Prod) + admin consent",
     "App regs created; admin consent granted on Graph permissions", 0),
    ("Compliance (Shell)", "CMP-6", "Service mailbox + Application Access Policy",
     "Service mailbox provisioned; access policy scoped to the app", 0),
    ("Compliance (Shell)", "CMP-7", "DNS hostnames + TLS certificates",
     "Prod/NonProd hostnames and TLS certs provisioned", 0),
]


def build():
    wb = Workbook()

    # ============================ 1. Overview ============================
    ws = wb.active
    ws.title = "Overview"
    title_block(ws, "VendorPulse — MVP Delivery Tracker",
                "Track build progress against the current in-scope MVP. Update daily/weekly. Keep the team focused; surface drift early.")
    ws["A4"] = "How to use this workbook"
    ws["A4"].font = Font(bold=True, size=11, color=NAVY)
    howto = [
        "1. The 'Task Tracker' is the main sheet — work the tasks, set Status and % Complete as you go.",
        "2. A module is only 'Done' when every row for it is Done AND its 'Module Exit Criteria' all pass.",
        "3. Log progress each day/week in 'Weekly Progress Log' to track velocity and catch drift early.",
        "4. The current scope is the MVP baseline ('MVP Scope'). Anything new from the discovery phase goes",
        "   into 'Scope Change Log' first — decide in/out before it touches the plan. Don't silently expand scope.",
        "5. 'Dashboard' auto-summarises status and % complete by area (formulas).",
        "6. Status values: Not Started, In Progress, In Review, Blocked, Done (colour-coded in Task Tracker).",
    ]
    r = 5
    for line in howto:
        ws.cell(row=r, column=1, value=line).font = TFONT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Legend — Status colours").font = Font(bold=True, size=11, color=NAVY)
    r += 1
    legend = [("Done", C_DONE), ("In Progress", C_PROG), ("In Review", C_REVIEW),
              ("Blocked", C_BLOCK), ("Not Started", C_NOT)]
    for name, col in legend:
        cell = ws.cell(row=r, column=1, value=name)
        cell.fill = PatternFill("solid", fgColor=col)
        cell.font = TFONT
        cell.border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Sheets").font = Font(bold=True, size=11, color=NAVY)
    r += 1
    idx = [
        "MVP Scope — what is in / out of the MVP baseline",
        "Milestones — phase/week plan with status",
        "Task Tracker — granular tasks + Definition of Done (main sheet)",
        "Module Exit Criteria — the gate checklist to call a module done",
        "Weekly Progress Log — running progress log",
        "Scope Change Log — discovery-phase changes vs the baseline",
        "Dashboard — auto summaries",
    ]
    for line in idx:
        ws.cell(row=r, column=1, value="• " + line).font = TFONT
        r += 1
    ws.column_dimensions["A"].width = 110

    # ============================ 2. MVP Scope ============================
    ws = wb.create_sheet("MVP Scope")
    title_block(ws, "MVP Scope Baseline",
                "This is the agreed MVP. Treat it as the frozen target. New asks go through Scope Change Log.")
    header_row(ws, 4, ["Category", "In scope (MVP)", "Out of scope / later"])
    set_widths(ws, [26, 60, 45])
    scope = [
        ("Workflow", "12-state forward-only WorkflowEngine; all 6 modules A–F", "Parallel/branching workflows; multi-cycle batch ops"),
        ("Module A — Scheduling", "Attendee refresh, availability, deterministic slot ranking, invite + Teams meeting after approval, RSVP", "Auto-rescheduling; external attendee self-service"),
        ("Module B — Scorecard", "Native in-app form, distribution via Graph mail, deterministic validation + compile + outliers", "Third-party survey tools; weighted scoring config UI"),
        ("Module C — Alignment", "AI draft action items + score-change summary, human-approved", "Auto-assignment of actions to owners"),
        ("Module D — Vendor Prep", "AI vendor brief + 3 response options per pushback, human-approved", "Vendor-facing portal"),
        ("Module E — Meeting", "Transcript parse, AI minutes, send after approval, Blob storage", "Live in-meeting transcription; real-time copilots"),
        ("Module F — Analytics", "Deterministic trends/trajectories + AI leadership brief card", "Predictive forecasting; cross-tenant benchmarking"),
        ("AI", "Single in-house AI Service -> Azure AI Foundry GPT-4o (in-tenant), draft text only, behind approval gate", "MAF SDK (removed); agentic multi-step autonomy; Anthropic/other LLMs"),
        ("Identity", "Entra SSO (OIDC) + RBAC (Coordinator/Sponsor/Viewer); Graph app-only certificate", "Guest/B2B external identities"),
        ("Hosting", "2 Azure VMs in Private VNet; App Gateway + WAF (Front Door removed)", "Container Apps / Foundry Hosted Agents (future)"),
        ("Data", "PostgreSQL Flexible Server (system of record); Blob for files; KV for secrets", "Data warehouse / BI export; pgvector RAG"),
        ("Observability", "App Insights + Log Analytics via OpenTelemetry; immutable audit", "Custom SIEM integration"),
        ("Integrations", "Microsoft Graph (Mail.Send, Calendars.ReadWrite, OnlineMeetings)", "Non-Microsoft mail/calendar"),
        ("Admin", "Admin module, Shell branding, Outlook templates", "Self-service tenant onboarding"),
    ]
    r = 5
    for row in scope:
        write_row(ws, r, row, wrap_cols=(2, 3))
        r += 1
    ws.freeze_panes = "A5"

    # ============================ 3. Milestones ============================
    ws = wb.create_sheet("Milestones")
    title_block(ws, "Milestones / Phase Plan",
                "High-level phases. Fill target dates and update status as you go.")
    header_row(ws, 4, ["Phase", "Key outcomes", "Target date", "Status", "% Complete", "Notes"])
    set_widths(ws, [30, 55, 14, 14, 12, 30])
    miles = [
        ("Phase 0 — Pre-mobilisation", "Shell provisions Entra app regs + admin consent, service mailbox, Azure infra, DNS/TLS, repo + CI/CD; LLM provider confirmed", "", "Not Started", 0, ""),
        ("Week 1 — Foundations & migration", "JSON->Postgres, Key Vault, remove Google, Entra SSO+RBAC, Graph cert, Module A live, design freeze", "", "Not Started", 0, ""),
        ("Week 2 — Development", "All modules built, scorecard, branding, Outlook templates, admin module, code freeze", "", "Not Started", 0, ""),
        ("Week 3 — Stabilisation & UAT", "Hardening, UAT by Shell VMO, observability + runbook, defect fixes, release readiness", "", "Not Started", 0, ""),
        ("Week 4 — Deployment & warranty", "IT Security sign-off, CAB, prod cutover & go-live, pilot cycle, training, warranty begins", "", "Not Started", 0, ""),
        ("Parallel — Compliance", "AI Registry/IAQ, EU AI Act classification, Shell.AI + TRB approval (external lead time)", "", "Not Started", 0, "Must clear before prod"),
    ]
    r = 5
    for row in miles:
        write_row(ws, r, row, wrap_cols=(2, 6))
        ws.cell(row=r, column=5).number_format = "0%"
        r += 1
    ws.freeze_panes = "A5"
    add_status_dv(ws, "D5", "D40")
    add_status_cf(ws, "D5:D40")

    # ============================ 4. Task Tracker ============================
    ws = wb.create_sheet("Task Tracker")
    title_block(ws, "Task Tracker",
                "The main sheet. Each row has a Definition of Done. Set Status + % as you progress. Filter by Area.")
    headers = ["Area", "ID", "Task / Deliverable", "Definition of Done (acceptance criteria)",
               "Owner", "Est (d)", "Status", "% Complete", "Start", "Target", "Actual done", "Notes / Blockers"]
    header_row(ws, 4, headers)
    set_widths(ws, [22, 8, 38, 50, 14, 7, 13, 11, 12, 12, 12, 30])
    r = 5
    last_area = None
    for area, tid, task, dod, est in TASKS:
        fill = AREAFILL if area != last_area else None
        write_row(ws, r, [area, tid, task, dod, "", est, "Not Started", 0, "", "", "", ""],
                  wrap_cols=(3, 4, 12))
        ws.cell(row=r, column=8).number_format = "0%"
        last_area = area
        r += 1
    last = r - 1
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:L{last}"
    add_status_dv(ws, "G5", f"G{last}")
    add_status_cf(ws, f"G5:G{last}")

    # ============================ 5. Module Exit Criteria ============================
    ws = wb.create_sheet("Module Exit Criteria")
    title_block(ws, "Module Exit Criteria (Definition of Done)",
                "A module is DONE only when ALL its criteria are met. Generic criteria apply to every module.")
    header_row(ws, 4, ["Scope", "Exit criterion (must be true to call it done)", "Met? (Y/N)", "Evidence / notes"])
    set_widths(ws, [22, 64, 12, 34])
    generic = [
        "State transitions enforced only via WorkflowEngine (invalid -> 409)",
        "Deterministic path works with ENABLE_LLM=false (where the module uses AI)",
        "All AI outputs pass through the approval gate — no auto-send",
        "AgentResponse contract honoured by both AI and deterministic paths",
        "Every run audited with correlation IDs (OTel -> App Insights)",
        "RBAC enforced server-side for the module's actions",
        "Unit + regression tests pass in CI",
        "No secrets in code; all via Key Vault + Managed Identity",
        "Module UAT scenario executed and signed off by Shell VMO",
    ]
    module_specific = {
        "Module A — Scheduling": [
            "Invite + Teams meeting created only after approval, from deterministic route",
            "Slot ranking is deterministic and reproducible (no LLM)",
            "Cycle reaches MEETING_SCHEDULED only through the valid state path",
        ],
        "Module B — Scorecard": [
            "All scorecards validated deterministically; no LLM in the data path",
            "Compile + outlier detection reproducible; SCORECARD_COMPILED reached",
        ],
        "Module C — Alignment": [
            "Action items + score-change summary are drafts, human-approved before use",
        ],
        "Module D — Vendor Prep": [
            "Brief + exactly 3 response options per pushback; approved before share",
        ],
        "Module E — Meeting": [
            "Minutes sent only after approval; transcript/minutes stored immutably in Blob",
            "POST_MEETING_COMPLETE reached only via the valid path",
        ],
        "Module F — Analytics": [
            "Analytics computed deterministically from history; brief card approved",
        ],
    }
    r = 5
    # generic block
    write_row(ws, r, ["ALL MODULES (generic)", "", "", ""], bold=True, fill=AREAFILL)
    r += 1
    for crit in generic:
        write_row(ws, r, ["", crit, "", ""], wrap_cols=(2, 4))
        r += 1
    for mod, crits in module_specific.items():
        write_row(ws, r, [mod, "", "", ""], bold=True, fill=AREAFILL)
        r += 1
        for crit in crits:
            write_row(ws, r, ["", crit, "", ""], wrap_cols=(2, 4))
            r += 1
    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(yn)
    yn.add(f"C5:C{r}")
    ws.freeze_panes = "A5"

    # ============================ 6. Weekly Progress Log ============================
    ws = wb.create_sheet("Weekly Progress Log")
    title_block(ws, "Weekly / Daily Progress Log",
                "One row per update. Track what shipped, overall %, blockers, and the next focus — to catch drift early.")
    header_row(ws, 4, ["Week / Period", "Date", "Focus (area/module)", "Done this period",
                       "% overall", "Blockers / risks", "Decisions", "Next focus", "Updated by"])
    set_widths(ws, [14, 12, 20, 40, 10, 28, 28, 28, 14])
    for r in range(5, 35):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
        ws.cell(row=r, column=5).number_format = "0%"
    ws.freeze_panes = "A5"

    # ============================ 7. Scope Change Log ============================
    ws = wb.create_sheet("Scope Change Log")
    title_block(ws, "Scope Change Log (vs MVP baseline)",
                "Discovery-phase or later changes land here first. Decide In MVP / Defer / Reject before touching the plan.")
    header_row(ws, 4, ["Change ID", "Date raised", "Raised by", "Description", "Type",
                       "In MVP?", "Impact (scope/time/cost)", "Decision", "Approved by", "Date decided"])
    set_widths(ws, [11, 12, 14, 40, 14, 12, 28, 20, 14, 12])
    for r in range(5, 35):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
    type_dv = DataValidation(type="list", formula1='"Add,Modify,Remove"', allow_blank=True)
    inmvp_dv = DataValidation(type="list", formula1='"Yes,No,Deferred"', allow_blank=True)
    ws.add_data_validation(type_dv)
    ws.add_data_validation(inmvp_dv)
    type_dv.add("E5:E40")
    inmvp_dv.add("F5:F40")
    ws.freeze_panes = "A5"

    # ============================ 8. Dashboard ============================
    ws = wb.create_sheet("Dashboard")
    title_block(ws, "Dashboard (auto)", "Summaries computed from the Task Tracker. Refreshes when you edit it.")
    tt = "'Task Tracker'"
    rng = f"{tt}!G5:G200"
    pct = f"{tt}!H5:H200"
    ws["A4"] = "Overall % complete"
    ws["A4"].font = Font(bold=True, size=11, color=NAVY)
    ws["B4"] = f"=IFERROR(AVERAGE({pct}),0)"
    ws["B4"].number_format = "0%"
    ws["B4"].font = Font(bold=True, size=14, color=RED)

    ws["A6"] = "Tasks by status"
    ws["A6"].font = Font(bold=True, size=11, color=NAVY)
    r = 7
    for s, col in [("Done", C_DONE), ("In Progress", C_PROG), ("In Review", C_REVIEW),
                   ("Blocked", C_BLOCK), ("Not Started", C_NOT)]:
        c1 = ws.cell(row=r, column=1, value=s)
        c1.fill = PatternFill("solid", fgColor=col)
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=f'=COUNTIF({rng},"{s}")')
        c2.border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="Total tasks").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f'=COUNTA({tt}!B5:B200)').font = Font(bold=True)

    ws["D6"] = "% complete by area"
    ws["D6"].font = Font(bold=True, size=11, color=NAVY)
    areas = []
    seen = set()
    for area, *_ in TASKS:
        if area not in seen:
            seen.add(area)
            areas.append(area)
    r = 7
    for area in areas:
        ws.cell(row=r, column=4, value=area).border = BORDER
        cell = ws.cell(row=r, column=5,
                       value=f'=IFERROR(AVERAGEIF({tt}!A5:A200,"{area}",{pct}),0)')
        cell.number_format = "0%"
        cell.border = BORDER
        r += 1
    set_widths(ws, [18, 12, 4, 26, 12])

    wb.save(OUT)
    print("Wrote", OUT)


def add_status_dv(ws, start, end):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{start}:{end}")


def add_status_cf(ws, rng):
    rules = [("Done", C_DONE), ("In Progress", C_PROG), ("In Review", C_REVIEW),
             ("Blocked", C_BLOCK), ("Not Started", C_NOT)]
    for val, col in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{val}"'],
                       fill=PatternFill("solid", fgColor=col)))


if __name__ == "__main__":
    build()
